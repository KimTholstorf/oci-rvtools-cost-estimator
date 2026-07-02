"""Anonymisation: sheet/column stripping, DNS + name tokenisation, cost parity."""

import csv

import pandas as pd
from openpyxl import load_workbook

from oci_rvtools.anonymize import anonymize_file, build_anonymized
from oci_rvtools.compute import aggregate_vinfo
from oci_rvtools.ingest import load_vinfo_dataframe


def _full_workbook(path):
    """A 'full' RVTools-like export: vInfo + vCPU + a sensitive extra sheet."""
    vinfo = pd.DataFrame({
        "VM": ["app1", "app2", "vCLS-xyz"],
        "Powerstate": ["poweredOn", "poweredOff", "poweredOn"],
        "Datacenter": ["Copenhagen-DC", "Copenhagen-DC", "Copenhagen-DC"],
        "Cluster": ["PROD-FIN", "PROD-FIN", "PROD-FIN"],
        "CPUs": [4, 8, 2],
        "Memory": [8192, 16384, 2048],
        "Total disk capacity MiB": [102400, 204800, 1024],
        "Provisioned MiB": [102400, 204800, 1024],
        "In Use MiB": [51200, 102400, 512],
        "DNS Name": ["app1.corp.example.com", "app2.corp.example.com", ""],
        "OS according to the VMware Tools": [
            "Oracle Linux 8 (64-bit)", "Microsoft Windows Server 2019 (64-bit)", ""],
        "OS according to the configuration file": ["", "", ""],
        # sensitive columns that must be dropped
        "Primary IP Address": ["10.1.1.5", "10.1.1.6", "10.1.1.7"],
        "Annotation": ["owner: alice", "billing: dept42", ""],
        "Folder": ["/DC/vm/finance", "/DC/vm/finance", "/DC/vm/system"],
    })
    vcpu = pd.DataFrame({
        "VM": ["app1", "app2", "vCLS-xyz"],
        "Sockets": [2, 4, 1],
        "OS according to the VMware Tools": [
            "Oracle Linux 8 (64-bit)", "Microsoft Windows Server 2019 (64-bit)", ""],
        "OS according to the configuration file": ["", "", ""],
    })
    vnet = pd.DataFrame({"VM": ["app1"], "IP Address": ["10.1.1.5"], "MAC Address": ["00:11:22:33:44:55"]})
    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        vinfo.to_excel(xl, sheet_name="vInfo", index=False)
        vcpu.to_excel(xl, sheet_name="vCPU", index=False)
        vnet.to_excel(xl, sheet_name="vNetwork", index=False)


def test_structure_and_stripping_no_names(tmp_path):
    src = tmp_path / "full.xlsx"
    _full_workbook(src)
    res = anonymize_file(src, anonymize_names=False)
    assert res is not None
    xlsx_path, key_path = res
    assert xlsx_path.name == "full_anonymized.xlsx"
    assert key_path is None  # no key file without --anonymize-names

    wb = load_workbook(xlsx_path)
    assert wb.sheetnames == ["vInfo", "vCPU"]  # vNetwork dropped

    vi = wb["vInfo"]
    headers = [c.value for c in vi[1]]
    assert "Primary IP Address" not in headers
    assert "Annotation" not in headers
    assert "Folder" not in headers
    assert "VM" in headers and "Cluster" in headers and "Datacenter" in headers

    # DNS domain stripped, names left real
    df = pd.read_excel(xlsx_path, sheet_name="vInfo")
    assert df.loc[df["VM"] == "app1", "DNS Name"].iloc[0] == "app1"
    assert set(df["VM"]) == {"app1", "app2", "vCLS-xyz"}
    assert set(df["Cluster"]) == {"PROD-FIN"}

    # vCPU stripped to VM + OS columns (Sockets dropped)
    vcpu_df = pd.read_excel(xlsx_path, sheet_name="vCPU")
    assert "Sockets" not in vcpu_df.columns
    assert "VM" in vcpu_df.columns


def test_name_tokenisation_and_key(tmp_path):
    src = tmp_path / "full.xlsx"
    _full_workbook(src)
    xlsx_path, key_path = anonymize_file(src, anonymize_names=True)
    assert key_path is not None and key_path.exists()

    vi = pd.read_excel(xlsx_path, sheet_name="vInfo")
    vc = pd.read_excel(xlsx_path, sheet_name="vCPU")

    # Real names gone; tokens present
    assert "app1" not in set(vi["VM"].astype(str))
    assert any(str(v).startswith("VM") for v in vi["VM"])
    # vCLS prefix preserved so cost pipeline still excludes it
    assert any(str(v).startswith("vCLS-") for v in vi["VM"])
    # Cluster / Datacenter tokenised
    assert all(str(c).startswith("cluster") for c in vi["Cluster"])
    assert all(str(d).startswith("dc") for d in vi["Datacenter"])
    # Hostname tokenised (blank stays blank)
    hosts = [str(h) for h in vi["DNS Name"] if pd.notna(h) and str(h).strip()]
    assert hosts and all(h.startswith("host") for h in hosts)

    # VM tokens consistent across vInfo and vCPU (join key intact)
    vi_map = dict(zip(pd.read_excel(src, sheet_name="vInfo")["VM"], vi["VM"]))
    vc_map = dict(zip(pd.read_excel(src, sheet_name="vCPU")["VM"], vc["VM"]))
    assert vi_map["app1"] == vc_map["app1"]
    assert vi_map["app2"] == vc_map["app2"]

    # Key file maps tokens back to originals
    with open(key_path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    cats = {r["Category"] for r in rows}
    assert {"VM", "Hostname", "Cluster", "Datacenter"} <= cats
    vm_rows = {r["Anonymized"]: r["Original"] for r in rows if r["Category"] == "VM"}
    assert "app1" in vm_rows.values() and "app2" in vm_rows.values()


def test_cost_parity_real_vs_anonymized(tmp_path):
    src = tmp_path / "full.xlsx"
    _full_workbook(src)
    xlsx_path, _ = anonymize_file(src, anonymize_names=True)

    real = load_vinfo_dataframe(src)
    anon = load_vinfo_dataframe(xlsx_path)
    for include_off in (True, False):
        r = aggregate_vinfo(real, include_vms_off=include_off, include_disks_off=True)
        a = aggregate_vinfo(anon, include_vms_off=include_off, include_disks_off=True)
        assert r == a, f"cost aggregation differs (include_off={include_off}): {r} vs {a}"


def test_structure_only_strips_vm_fqdn_and_keeps_join(tmp_path):
    """Structure-only mode strips FQDN domains from VM names in BOTH sheets so
    the vInfo<->vCPU OS-fallback join still matches."""
    src = tmp_path / "fqdn.xlsx"
    vinfo = pd.DataFrame({
        "VM": ["web01.corp.local", "db02.corp.local", "plain-name"],
        "Powerstate": ["poweredOn", "poweredOn", "poweredOn"],
        "CPUs": [4, 8, 2],
        "Memory": [8192, 16384, 4096],
        "Total disk capacity MiB": [102400, 204800, 51200],
        "Provisioned MiB": [102400, 204800, 51200],
        "In Use MiB": [51200, 102400, 25600],
        "DNS Name": ["web01.corp.local", "db02.corp.local", "plain-name"],
        # OS deliberately absent from vInfo -> must come from vCPU via the join
    })
    vcpu = pd.DataFrame({
        "VM": ["web01.corp.local", "db02.corp.local", "plain-name"],
        "OS according to the VMware Tools": [
            "Ubuntu Linux (64-bit)", "Oracle Linux 8 (64-bit)", "CentOS 7 (64-bit)"],
        "OS according to the configuration file": ["", "", ""],
    })
    with pd.ExcelWriter(src, engine="openpyxl") as xl:
        vinfo.to_excel(xl, sheet_name="vInfo", index=False)
        vcpu.to_excel(xl, sheet_name="vCPU", index=False)

    xlsx_path, key_path = anonymize_file(src, anonymize_names=False)
    assert key_path is None

    vi = pd.read_excel(xlsx_path, sheet_name="vInfo")
    vc = pd.read_excel(xlsx_path, sheet_name="vCPU")
    assert set(vi["VM"]) == {"web01", "db02", "plain-name"}
    assert set(vc["VM"]) == {"web01", "db02", "plain-name"}

    # The OS-fallback join must still work on the stripped names
    anon = load_vinfo_dataframe(xlsx_path)
    os_by_vm = anon.set_index("VM")["OS according to the VMware Tools"].to_dict()
    assert os_by_vm["web01"] == "Ubuntu Linux (64-bit)"
    assert os_by_vm["db02"] == "Oracle Linux 8 (64-bit)"


def test_missing_vinfo_returns_none(tmp_path):
    src = tmp_path / "novinfo.xlsx"
    pd.DataFrame({"x": [1]}).to_excel(src, sheet_name="vNetwork", index=False)
    assert anonymize_file(src, anonymize_names=False) is None


def test_cli_anonymize_full(tmp_path):
    """`--anonymize-full` tokenises names and emits a key file."""
    from oci_rvtools.cli import main

    src = tmp_path / "full.xlsx"
    _full_workbook(src)
    rc = main(["--anonymize-full", "--rvtools", str(src)])
    assert rc == 0
    assert (tmp_path / "full_anonymized.xlsx").exists()
    assert (tmp_path / "full_anonymized_key.csv").exists()

    vi = pd.read_excel(tmp_path / "full_anonymized.xlsx", sheet_name="vInfo")
    assert any(str(v).startswith("VM") for v in vi["VM"])


def test_cli_anonymize(tmp_path):
    """`--anonymize` keeps real names and writes no key file."""
    from oci_rvtools.cli import main

    src = tmp_path / "full.xlsx"
    _full_workbook(src)
    rc = main(["--anonymize", "--rvtools", str(src)])
    assert rc == 0
    assert (tmp_path / "full_anonymized.xlsx").exists()
    assert not (tmp_path / "full_anonymized_key.csv").exists()

    vi = pd.read_excel(tmp_path / "full_anonymized.xlsx", sheet_name="vInfo")
    assert "app1" in set(vi["VM"].astype(str))  # names kept real


def test_cli_anonymize_flags_mutually_exclusive(tmp_path):
    """Passing both --anonymize and --anonymize-full is rejected by argparse."""
    import pytest
    from oci_rvtools.cli import main

    src = tmp_path / "full.xlsx"
    _full_workbook(src)
    with pytest.raises(SystemExit):
        main(["--anonymize", "--anonymize-full", "--rvtools", str(src)])
