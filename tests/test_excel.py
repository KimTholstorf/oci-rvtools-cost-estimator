"""End-to-end workbook smoke test — guards the column registry cross-sheet refs."""

import pandas as pd
from openpyxl import load_workbook

from oci_rvtools.model import LineItem
from oci_rvtools.report import write_output
from oci_rvtools.report import columns


def _line(cat, name, part, raw, usage, price):
    return LineItem(description=name, part_number=part, category=cat,
                    raw_base_quantity=raw, usage_quantity=usage, unit_price=price)


def _sheets():
    def make():
        return [
            _line("ocpu", "OCPU", "B97384", 4.0, 730, 0.025),
            _line("memory", "Memory", "B97385", 10.0, 730, 0.0015),
            _line("storage", "Storage", "B91961", 100.0, 1.0, 0.0255),
            _line("vpu", "VPU", "B91962", 1000.0, 1.0, 0.0017),
        ]
    return {"total_disk": make(), "used_disk": make()}


def _metadata():
    return {
        "source_files": "test.xlsx",
        "filter_datacenters": [],
        "filter_clusters": [],
        "hours_per_month": "730",
        "currency": "USD",
        "vpu": "10.0",
        "powered_on_vms": 2,
        "powered_off_vms": 1,
        "powered_off_included": False,
        "powered_off_disks_included": True,
    }


def _vm_df():
    return pd.DataFrame({
        "VM": ["app1", "winsrv", "old"],
        "Powerstate": ["poweredOn", "poweredOn", "poweredOff"],
        "CPUs": [4, 8, 2],
        "Memory": [8192, 16384, 4096],
        "Total disk capacity MiB": [102400, 204800, 51200],
        "Provisioned MiB": [102400, 204800, 51200],
        "In Use MiB": [51200, 102400, 25600],
        "OS according to the VMware Tools": [
            "Oracle Linux 8 (64-bit)",
            "Microsoft Windows Server 2019 (64-bit)",
            "Microsoft Windows Server 2008 R2 (64-bit)",
        ],
        "OS according to the configuration file": ["", "", ""],
    })


def test_workbook_structure_and_cross_sheet_refs(tmp_path):
    out = tmp_path / "out.xlsx"
    vm_df = _vm_df()
    write_output(out, _metadata(), _sheets(), "USD", vm_df=vm_df)

    wb = load_workbook(out)
    assert wb.sheetnames == ["Cost Summary", "VM Details", "OS Summary"]

    cs = wb["Cost Summary"]
    # OCPU line item lives at row 13 (8 metadata rows + layout offsets)
    assert cs["A13"].value == "OCPU"
    end = columns.vd_data_end(len(vm_df))
    # This metadata excludes powered-off VMs, so OCPU Part Qty is a powered-on SUMIF over col E
    assert cs["C13"].value == columns.vd_sum_powered_on("ocpu", end)
    # Monthly cost formula + accounting format
    assert cs["G13"].value == "=C13*D13*E13*F13"
    assert "#,##0.00" in cs["G13"].number_format
    # Powered-off disks ARE included here, so used-disk storage Part Qty is a plain SUM over col H
    storage_rows = [r for r in range(1, cs.max_row + 1) if cs.cell(row=r, column=1).value == "Storage"]
    assert len(storage_rows) == 2
    assert cs.cell(row=storage_rows[1], column=3).value == columns.vd_sum("disk_used_gb", end)


def test_poweredoff_flag_changes_ocpu_part_qty(tmp_path):
    """Regression: the include/exclude powered-off VMs toggle must change the
    Cost Summary OCPU Part Qty (SUMIF powered-on vs SUM all)."""
    vm_df = _vm_df()
    end = columns.vd_data_end(len(vm_df))

    md_excl = _metadata()
    md_excl["powered_off_included"] = False
    out_excl = tmp_path / "excl.xlsx"
    write_output(out_excl, md_excl, _sheets(), "USD", vm_df=vm_df)

    md_incl = _metadata()
    md_incl["powered_off_included"] = True
    out_incl = tmp_path / "incl.xlsx"
    write_output(out_incl, md_incl, _sheets(), "USD", vm_df=vm_df)

    c_excl = load_workbook(out_excl)["Cost Summary"]["C13"].value
    c_incl = load_workbook(out_incl)["Cost Summary"]["C13"].value

    assert c_excl == columns.vd_sum_powered_on("ocpu", end)
    assert c_incl == columns.vd_sum("ocpu", end)
    assert c_excl != c_incl, "powered-off toggle had no effect on OCPU Part Qty"


def test_vm_details_headers_and_conditional_formatting(tmp_path):
    out = tmp_path / "out.xlsx"
    write_output(out, _metadata(), _sheets(), "USD", vm_df=_vm_df())

    wb = load_workbook(out)
    vd = wb["VM Details"]
    assert vd["C1"].value == "OS Detected"
    assert vd["D1"].value == "OCI Compatible"
    assert vd["E1"].value == "OCPU"
    # Row 3 = first VM (sorted: app1) -> Oracle Linux -> yes
    assert vd["C3"].value == "Oracle Linux 8 (64-bit)"
    assert vd["D3"].value == "yes"
    # Conditional formatting registered on the sheet
    assert len(list(vd.conditional_formatting)) >= 1


def test_os_summary_present(tmp_path):
    out = tmp_path / "out.xlsx"
    write_output(out, _metadata(), _sheets(), "USD", vm_df=_vm_df())

    wb = load_workbook(out)
    os_sheet = wb["OS Summary"]
    all_formulas = [
        os_sheet.cell(row=r, column=c).value
        for r in range(1, os_sheet.max_row + 1)
        for c in range(1, 4)
        if isinstance(os_sheet.cell(row=r, column=c).value, str)
    ]
    assert any(f.startswith("=COUNTIF") for f in all_formulas)


def test_accounting_format_non_usd(tmp_path):
    out = tmp_path / "dkk.xlsx"
    md = _metadata()
    md["currency"] = "DKK"
    write_output(out, md, _sheets(), "DKK", vm_df=_vm_df())
    wb = load_workbook(out)
    cs = wb["Cost Summary"]
    assert "[$DKK]" in cs["G13"].number_format
