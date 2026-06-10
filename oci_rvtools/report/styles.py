# Copyright (c) 2026 Kim Tholstorf
# https://github.com/KimTholstorf/oci-rvtools-cost-estimator
# MIT License — see LICENSE file for details

"""Shared Excel visual primitives: fonts, fills, row heights, number formats."""

from __future__ import annotations

from typing import Dict

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill

# Fonts / fills / alignment
TITLE_FONT = Font(bold=True, size=14)
HEADER_FILL = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
SECTION_LABEL_FILL = PatternFill(start_color="FFB8CCE4", end_color="FFB8CCE4", fill_type="solid")
DISCLAIMER_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

# Row heights
TITLE_ROW_HEIGHT = 40
HEADER_ROW_HEIGHT = 40
DEFAULT_ROW_HEIGHT = 20
DISCLAIMER_ROW_HEIGHT = 80

# OCI-compatibility conditional-formatting fills — applied as Excel rules so the
# colour follows the cell value if it is edited.
CF_FILL_YES = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")    # soft green
CF_FILL_MAYBE = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")  # soft amber
CF_FILL_NO = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")     # soft red


def add_compat_rules(ws, cell_range: str) -> None:
    """Attach yes/maybe/no colour rules to a range of OCI-compatibility cells."""
    ws.conditional_formatting.add(cell_range, CellIsRule(operator="equal", formula=['"yes"'], fill=CF_FILL_YES))
    ws.conditional_formatting.add(cell_range, CellIsRule(operator="equal", formula=['"maybe"'], fill=CF_FILL_MAYBE))
    ws.conditional_formatting.add(cell_range, CellIsRule(operator="equal", formula=['"no"'], fill=CF_FILL_NO))


# Currencies with a well-known single glyph; everything else uses [$CODE].
_CURRENCY_SYMBOLS: Dict[str, str] = {
    "USD": "$", "EUR": "€", "GBP": "£",
    "JPY": "¥", "CNY": "¥", "INR": "₹", "KRW": "₩",
}


def accounting_number_format(currency: str) -> str:
    """Return an Excel accounting number-format string for an ISO currency code.

    Matches what Excel writes when you pick Accounting and choose the currency:
    a recognised glyph where one exists, otherwise the ISO code in ``[$CODE]``
    notation displayed left-aligned.
    """
    code = currency.upper()
    sym = _CURRENCY_SYMBOLS.get(code, f"[${code}]")
    return f'_({sym}* #,##0.00_);_({sym}* (#,##0.00);_({sym}* "-"??_);_(@_)'
