# Copyright (c) 2026 Kim Tholstorf
# https://github.com/KimTholstorf/oci-rvtools-cost-estimator
# MIT License — see LICENSE file for details

"""The VM Details sheet: one row per VM with per-VM cost formulas."""

from __future__ import annotations

import math
from typing import Dict

import pandas as pd
from openpyxl.styles import Alignment, Font

from ..compute import MIB_TO_GB
from ..osmatch import detect_os
from . import columns
from .columns import col_index, col_letter, vd_local_ref
from .cost_summary import COST_SUMMARY_SHEET, cost_summary_layout
from .styles import (
    DEFAULT_ROW_HEIGHT,
    HEADER_FILL,
    HEADER_ROW_HEIGHT,
    add_compat_rules,
)


def write_vm_detail_sheet(
    wb,
    vm_df: pd.DataFrame,
    metadata: Dict[str, object],
    currency: str,
    include_vms_off: bool,
    include_disks_off: bool,
) -> None:
    """Append a 'VM Details' sheet with one row per VM and per-VM cost formulas."""
    if vm_df is None or vm_df.empty:
        return

    ws = wb.create_sheet(title=columns.VM_DETAILS_SHEET)

    # Cross-sheet references into Cost Summary (unit prices in col F, meta in col B).
    layout = cost_summary_layout(metadata)
    cs = COST_SUMMARY_SHEET
    hours_ref = f"'{cs}'!B${layout['hours_row']}"
    ocpu_price = f"'{cs}'!F${layout['ocpu_row']}"
    ram_price = f"'{cs}'!F${layout['ram_row']}"
    stor_price = f"'{cs}'!F${layout['stor_row']}"
    vpu_pgb = f"'{cs}'!B${layout['vpu_meta_row']}"
    vpu_price = f"'{cs}'!F${layout['vpu_row']}"
    disk_cost_per_gb = f"({stor_price}+{vpu_pgb}*{vpu_price})"

    # Column widths
    for letter, width in columns.widths().items():
        ws.column_dimensions[letter].width = width

    n_cols = columns.column_count()

    # Row 1: headers
    for col_idx, header in enumerate(columns.headers(currency), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = HEADER_FILL
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    # Row 2: rounding note
    note_cell = ws.cell(row=2, column=1, value=columns.VM_DETAIL_NOTE)
    note_cell.font = Font(italic=True, color="00888888")
    note_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[2].height = 36

    df = vm_df.sort_values("VM", key=lambda s: s.str.lower()).reset_index(drop=True)

    def setc(row: int, key: str, value) -> None:
        ws.cell(row=row, column=col_index(key), value=value)

    for idx, row_data in df.iterrows():
        r = int(idx) + columns.VM_DETAILS_DATA_START

        powerstate = str(row_data["Powerstate"])
        is_powered_off = powerstate.strip().lower() != "poweredon"

        cpus = float(pd.to_numeric(row_data["CPUs"], errors="coerce") or 0)
        mem_mib = float(pd.to_numeric(row_data["Memory"], errors="coerce") or 0)
        total_mib = float(pd.to_numeric(row_data["Total disk capacity MiB"], errors="coerce") or 0)
        prov_mib = float(pd.to_numeric(row_data["Provisioned MiB"], errors="coerce") or 0)
        used_mib = float(pd.to_numeric(row_data["In Use MiB"], errors="coerce") or 0)

        eff_prov_mib = total_mib if total_mib > 0 else prov_mib

        ocpu_val = math.ceil(cpus / 2.0)
        ram_gb_val = math.ceil(mem_mib / 1024.0)
        disk_prov_val = math.ceil(eff_prov_mib * MIB_TO_GB)
        disk_used_val = math.ceil(used_mib * MIB_TO_GB)

        include_cpu_ram = not is_powered_off or include_vms_off
        include_disk = not is_powered_off or include_disks_off

        os_detected, oci_compat, compat_note = detect_os(row_data)

        # Descriptive + quantity columns
        setc(r, "vm", str(row_data["VM"]))
        setc(r, "powerstate", powerstate)
        setc(r, "os_detected", os_detected)
        setc(r, "oci_compatible", oci_compat)
        setc(r, "ocpu", ocpu_val)
        setc(r, "ram_gb", ram_gb_val)
        setc(r, "disk_prov_gb", disk_prov_val)
        setc(r, "disk_used_gb", disk_used_val)

        # Cost formulas (or 0 when excluded by power-state flags)
        if include_cpu_ram:
            setc(r, "ocpu_cost", f"={vd_local_ref('ocpu', r)}*{hours_ref}*{ocpu_price}")
            setc(r, "ram_cost", f"={vd_local_ref('ram_gb', r)}*{hours_ref}*{ram_price}")
        else:
            setc(r, "ocpu_cost", 0)
            setc(r, "ram_cost", 0)

        if include_disk:
            setc(r, "disk_prov_cost", f"={vd_local_ref('disk_prov_gb', r)}*{disk_cost_per_gb}")
            setc(r, "disk_used_cost", f"={vd_local_ref('disk_used_gb', r)}*{disk_cost_per_gb}")
        else:
            setc(r, "disk_prov_cost", 0)
            setc(r, "disk_used_cost", 0)

        # Totals
        ocpu_c = vd_local_ref("ocpu_cost", r)
        ram_c = vd_local_ref("ram_cost", r)
        prov_c = vd_local_ref("disk_prov_cost", r)
        used_c = vd_local_ref("disk_used_cost", r)
        setc(r, "monthly_prov", f"={ocpu_c}+{ram_c}+{prov_c}")
        setc(r, "monthly_used", f"={ocpu_c}+{ram_c}+{used_c}")
        setc(r, "yearly_prov", f"={vd_local_ref('monthly_prov', r)}*12")
        setc(r, "yearly_used", f"={vd_local_ref('monthly_used', r)}*12")

        # Note — combine powerstate and OS compat notes where applicable
        note_parts = []
        if is_powered_off:
            note_parts.append("Powered Off")
        if compat_note:
            note_parts.append(compat_note)
        setc(r, "note", " | ".join(note_parts))

        ws.row_dimensions[r].height = DEFAULT_ROW_HEIGHT

    # Conditional formatting for the OCI Compatible column
    last_data_row = columns.vd_data_end(len(df))
    compat_col = col_letter("oci_compatible")
    add_compat_rules(ws, f"{compat_col}{columns.VM_DETAILS_DATA_START}:{compat_col}{last_data_row}")
