# Copyright (c) 2026 Kim Tholstorf
# https://github.com/KimTholstorf/oci-rvtools-cost-estimator
# MIT License — see LICENSE file for details

"""VM Details column registry — the single source of truth for that sheet's layout.

Column order, headers, widths, and every cross-sheet reference are derived from
``VM_DETAIL_SCHEMA``. Inserting or reordering a column is a one-line edit here;
the writer and the Cost Summary / OS Summary formulas pick it up automatically
via ``col_index`` / ``col_letter`` / ``vd_*`` helpers — no manual cell shuffling.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List

from openpyxl.utils import get_column_letter

VM_DETAILS_SHEET = "VM Details"
# Row 1 = headers, row 2 = note, row 3+ = data.
VM_DETAILS_DATA_START = 3


@dataclass(frozen=True)
class VMColumn:
    key: str          # stable identifier used by writers and cross-sheet refs
    header: str       # display header (may contain a {currency} placeholder)
    width: float      # Excel column width (character units)


VM_DETAIL_SCHEMA: List[VMColumn] = [
    VMColumn("vm",             "VM Name",                              34),
    VMColumn("powerstate",     "Powerstate",                          14),
    VMColumn("os_detected",    "OS Detected",                       37.1),
    VMColumn("oci_compatible", "OCI Compatible",                      16),
    VMColumn("ocpu",           "OCPU",                                10),
    VMColumn("ram_gb",         "RAM (GB)",                            12),
    VMColumn("disk_prov_gb",   "Disk Provisioned (GB)",               22),
    VMColumn("disk_used_gb",   "Disk Used (GB)",                      18),
    VMColumn("ocpu_cost",      "OCPU Cost ({currency})",              22),
    VMColumn("ram_cost",       "RAM Cost ({currency})",               22),
    VMColumn("disk_prov_cost", "Disk Prov Cost ({currency})",         24),
    VMColumn("disk_used_cost", "Disk Used Cost ({currency})",         24),
    VMColumn("monthly_prov",   "Total Monthly - Prov ({currency})",   28),
    VMColumn("monthly_used",   "Total Monthly - Used ({currency})",   26),
    VMColumn("yearly_prov",    "Total Yearly - Prov ({currency})",    28),
    VMColumn("yearly_used",    "Total Yearly - Used ({currency})",    26),
    VMColumn("note",           "Note",                                16),
]

_INDEX: Dict[str, int] = {col.key: i for i, col in enumerate(VM_DETAIL_SCHEMA)}

VM_DETAIL_NOTE = (
    "ⓘ  The aggregate totals on the Cost Summary sheet are calculated by summing "
    "the per-VM values in this sheet. Editing a value here will automatically "
    "update the monthly cost totals."
)


def col_index(key: str) -> int:
    """1-based column index for a schema key."""
    return _INDEX[key] + 1


def col_letter(key: str) -> str:
    """Excel column letter for a schema key (e.g. "ocpu" -> "E")."""
    return get_column_letter(col_index(key))


def headers(currency: str) -> List[str]:
    return [col.header.format(currency=currency) for col in VM_DETAIL_SCHEMA]


def widths() -> Dict[str, float]:
    return {get_column_letter(i + 1): col.width for i, col in enumerate(VM_DETAIL_SCHEMA)}


def column_count() -> int:
    return len(VM_DETAIL_SCHEMA)


def vd_local_ref(key: str, row: int) -> str:
    """Same-sheet reference, e.g. ``E5``."""
    return f"{col_letter(key)}{row}"


def vd_abs_ref(key: str, row: int) -> str:
    """Absolute cross-sheet reference, e.g. ``'VM Details'!$E$5``."""
    return f"'{VM_DETAILS_SHEET}'!${col_letter(key)}${row}"


def vd_sum(key: str, end_row: int) -> str:
    """SUM over a column's data rows, e.g. ``=SUM('VM Details'!E3:E690)``."""
    letter = col_letter(key)
    return f"=SUM('{VM_DETAILS_SHEET}'!{letter}{VM_DETAILS_DATA_START}:{letter}{end_row})"


def vd_data_end(n_rows: int) -> int:
    """Last data row for ``n_rows`` VMs."""
    return VM_DETAILS_DATA_START - 1 + n_rows
