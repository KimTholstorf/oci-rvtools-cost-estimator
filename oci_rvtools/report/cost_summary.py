# Copyright (c) 2026 Kim Tholstorf
# https://github.com/KimTholstorf/oci-rvtools-cost-estimator
# MIT License — see LICENSE file for details

"""The Cost Summary sheet: metadata block plus two priced disk sections."""

from __future__ import annotations

from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font

from ..model import LineItem
from . import columns
from .styles import (
    DEFAULT_ROW_HEIGHT,
    DISCLAIMER_ALIGNMENT,
    DISCLAIMER_ROW_HEIGHT,
    HEADER_FILL,
    HEADER_ROW_HEIGHT,
    SECTION_LABEL_FILL,
    TITLE_FONT,
    TITLE_ROW_HEIGHT,
    accounting_number_format,
)

COST_SUMMARY_SHEET = "Cost Summary"

TABLE_COLUMN_COUNT = 7
EXCEL_HEADERS = [
    "Description",
    "Part Number",
    "Part Qty",
    "Instance Qty",
    "Usage Qty",
    "Unit Price ({currency})",
    "Monthly Cost ({currency})",
]
COLUMN_WIDTHS = {"A": 36, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16}

SECTION_LABELS = {
    "total_disk": "Total Provisioned Disk",
    "used_disk": "Total Used Disk",
}


def build_metadata_rows(metadata: Dict[str, object]) -> List[Tuple[str, str, str]]:
    filter_parts = []
    if metadata.get("filter_datacenters"):
        filter_parts.append(f"Datacenter: {', '.join(metadata['filter_datacenters'])}")
    if metadata.get("filter_clusters"):
        filter_parts.append(f"Cluster: {', '.join(metadata['filter_clusters'])}")
    filter_value = " | ".join(filter_parts) if filter_parts else "None"

    return [
        ("Source Files", metadata["source_files"], ""),
        ("Filters", filter_value, ""),
        ("Hours per Month", metadata["hours_per_month"], ""),
        ("Currency", metadata["currency"], ""),
        ("VPU", metadata["vpu"], ""),
        ("Powered On VMs", str(metadata["powered_on_vms"]), "(included)"),
        (
            "Powered Off VMs",
            str(metadata["powered_off_vms"]),
            "(included)" if metadata["powered_off_included"] else "(excluded)",
        ),
        (
            "Powered Off Disks",
            "",
            "(included)" if metadata["powered_off_disks_included"] else "(excluded)",
        ),
    ]


def cost_summary_layout(metadata: Dict[str, object]) -> Dict[str, int]:
    """Row coordinates of key Cost Summary cells, derived from the metadata block.

    Returned keys: ``hours_row``, ``vpu_meta_row`` (metadata values in col B) and
    ``ocpu_row``/``ram_row``/``stor_row``/``vpu_row`` (line-item rows; unit price
    in col F). Centralising this lets VM Details reference Cost Summary without
    hardcoding row numbers.
    """
    meta_rows = build_metadata_rows(metadata)
    n_meta = len(meta_rows)
    hours_row = 2 + next(i for i, (k, _, _) in enumerate(meta_rows) if k == "Hours per Month")
    vpu_meta_row = 2 + next(i for i, (k, _, _) in enumerate(meta_rows) if k == "VPU")
    data_start = 2 + n_meta + 3  # title(1) + metadata + blank + section_label + header
    return {
        "hours_row": hours_row,
        "vpu_meta_row": vpu_meta_row,
        "ocpu_row": data_start,
        "ram_row": data_start + 1,
        "stor_row": data_start + 2,
        "vpu_row": data_start + 3,
    }


def _format_row(ws, row: int, *, bold: bool = False, vertical: str = "bottom") -> None:
    for col_idx in range(1, TABLE_COLUMN_COUNT + 1):
        cell = ws.cell(row=row, column=col_idx)
        if bold:
            cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical=vertical)


def _append_advisory(ws, total_row: int) -> Tuple[int, int]:
    quote_row = total_row + 3
    quote_cell = ws.cell(row=quote_row, column=1, value="Quote is for investment proposal only.")
    quote_cell.font = Font(bold=True, color="00FF0000")
    ws.merge_cells(start_row=quote_row, start_column=1, end_row=quote_row, end_column=TABLE_COLUMN_COUNT)

    disclaimer_row = quote_row + 2
    disclaimer_text = (
        "Disclaimer:  This sample quote is provided solely for evaluation purposes and is intended to further "
        "discussions between you and Oracle.  This sample quote is not eligible for acceptance by you and is "
        "not a binding contract between you and Oracle for the services specified.  If you would like to purchase "
        "the services specified in this sample quote, please request that Oracle issue you a formal quote (which "
        "may include an OMA or a CSA if you do not already have an appropriate agreement in place with Oracle) "
        "for your acceptance and execution.  Your formal quote will be effective only upon Oracle's acceptance "
        "of the formal quote (and the OMA or CSA, if required)."
    )
    disclaimer_cell = ws.cell(row=disclaimer_row, column=1, value=disclaimer_text)
    ws.merge_cells(start_row=disclaimer_row, start_column=1, end_row=disclaimer_row, end_column=TABLE_COLUMN_COUNT)
    disclaimer_cell.alignment = DISCLAIMER_ALIGNMENT
    ws.row_dimensions[disclaimer_row].height = DISCLAIMER_ROW_HEIGHT
    return quote_row, disclaimer_row


def write_cost_summary(
    ws,
    metadata: Dict[str, object],
    sheets: Dict[str, List[LineItem]],
    currency: str,
    vm_df: Optional[pd.DataFrame] = None,
) -> None:
    """Render the Cost Summary sheet onto ``ws``."""
    headers = [h.format(currency=currency) for h in EXCEL_HEADERS]
    cost_fmt = accounting_number_format(currency)
    layout = cost_summary_layout(metadata)

    def roundup_formula(value: float) -> str:
        raw_str = f"{value:.10f}".rstrip("0").rstrip(".")
        if not raw_str:
            raw_str = "0"
        return f"=ROUNDUP({raw_str},0)"

    use_vd_sums = vm_df is not None and not vm_df.empty
    vd_end = columns.vd_data_end(len(vm_df)) if use_vd_sums else None

    # Part Qty for OCPU/RAM/Storage is the SUM of the corresponding VM Details column.
    qty_column_keys = {"ocpu": "ocpu", "memory": "ram_gb"}

    def part_qty_formula(category: str, sheet_key: str, raw_quantity: float) -> str:
        if not use_vd_sums:
            return roundup_formula(raw_quantity)
        if category == "storage":
            key = "disk_prov_gb" if sheet_key == "total_disk" else "disk_used_gb"
        else:
            key = qty_column_keys.get(category)
        if key is None:
            return roundup_formula(raw_quantity)
        return columns.vd_sum(key, vd_end)

    # Column widths
    for column, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width

    # Row 1: title banner
    title_text = f"Oracle Investment Proposal (as of {datetime.now().strftime('%m/%d/%Y')})"
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TABLE_COLUMN_COUNT)
    ws.row_dimensions[1].height = TITLE_ROW_HEIGHT

    # Metadata block (rows 2+)
    meta_rows = build_metadata_rows(metadata)
    row = 2
    for title, value, status in meta_rows:
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=2, value=value)
        if status:
            ws.cell(row=row, column=3, value=status)
        ws.row_dimensions[row].height = DEFAULT_ROW_HEIGHT
        row += 1
    metadata_end_row = row - 1
    for r in range(2, metadata_end_row + 1):
        _format_row(ws, r)
    _format_row(ws, 2, bold=True)

    # Two priced sections (total_disk then used_disk)
    cursor = metadata_end_row + 2
    last_total_row = None

    for sheet_key, items in sheets.items():
        section_label_row = cursor
        lbl = ws.cell(row=section_label_row, column=1, value=SECTION_LABELS.get(sheet_key, sheet_key))
        lbl.font = Font(bold=True)
        lbl.fill = SECTION_LABEL_FILL
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(
            start_row=section_label_row, start_column=1,
            end_row=section_label_row, end_column=TABLE_COLUMN_COUNT,
        )
        ws.row_dimensions[section_label_row].height = DEFAULT_ROW_HEIGHT

        header_row = section_label_row + 1
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.fill = HEADER_FILL
        ws.row_dimensions[header_row].height = HEADER_ROW_HEIGHT

        data_start = header_row + 1
        current_row = data_start
        storage_row_index: Optional[int] = None

        for item in items:
            ws.cell(row=current_row, column=1, value=item.description)
            ws.cell(row=current_row, column=2, value=item.part_number)

            part_cell = ws.cell(row=current_row, column=3)
            instance_cell = ws.cell(row=current_row, column=4)
            usage_cell = ws.cell(row=current_row, column=5)

            if item.category == "vpu":
                if storage_row_index is None:
                    raise ValueError("Storage row must be written before VPU row.")
                part_cell.value = f"=C{storage_row_index}*B${layout['vpu_meta_row']}"
                instance_cell.value = 1.0
                usage_cell.value = 1.0
            else:
                part_cell.value = part_qty_formula(item.category, sheet_key, item.raw_base_quantity)
                instance_cell.value = 1.0
                if item.category in {"ocpu", "memory"}:
                    usage_cell.value = f"=B${layout['hours_row']}"
                elif item.category == "storage":
                    usage_cell.value = 1.0
                    storage_row_index = current_row
                else:
                    usage_cell.value = item.usage_quantity

            ws.cell(row=current_row, column=6, value=item.unit_price)
            cost_cell = ws.cell(
                row=current_row, column=7,
                value=f"=C{current_row}*D{current_row}*E{current_row}*F{current_row}",
            )
            cost_cell.number_format = cost_fmt
            ws.row_dimensions[current_row].height = DEFAULT_ROW_HEIGHT
            current_row += 1

        total_row = current_row
        ws.cell(row=total_row, column=1, value="Monthly Total")
        for col in range(2, 7):
            ws.cell(row=total_row, column=col, value="")
        total_cost_cell = ws.cell(row=total_row, column=7, value=f"=SUM(G{data_start}:G{current_row - 1})")
        total_cost_cell.number_format = cost_fmt
        ws.row_dimensions[total_row].height = DEFAULT_ROW_HEIGHT

        for r in range(data_start, total_row):
            _format_row(ws, r)
        _format_row(ws, total_row, bold=True)

        last_total_row = total_row
        cursor = total_row + 2

    _, disclaimer_row = _append_advisory(ws, last_total_row)
    ws.row_dimensions[disclaimer_row].height = DISCLAIMER_ROW_HEIGHT
