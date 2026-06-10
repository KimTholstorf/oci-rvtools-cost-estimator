# Copyright (c) 2026 Kim Tholstorf
# https://github.com/KimTholstorf/oci-rvtools-cost-estimator
# MIT License — see LICENSE file for details

"""The OS Summary sheet: compatibility overview and per-OS breakdown.

All counts are COUNTIF formulas referencing VM Details, so the summary stays in
sync when VM Details cells are edited.
"""

from __future__ import annotations

from typing import Dict, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font

from ..osmatch import detect_os
from . import columns
from .columns import col_letter
from .cost_summary import COST_SUMMARY_SHEET  # noqa: F401  (kept for symmetry)
from .styles import (
    DEFAULT_ROW_HEIGHT,
    HEADER_FILL,
    HEADER_ROW_HEIGHT,
    SECTION_LABEL_FILL,
    TITLE_FONT,
    TITLE_ROW_HEIGHT,
    add_compat_rules,
)

OS_SUMMARY_SHEET = "OS Summary"


def _vd_abs_range(key: str, end_row: int) -> str:
    letter = col_letter(key)
    return (
        f"'{columns.VM_DETAILS_SHEET}'!${letter}${columns.VM_DETAILS_DATA_START}"
        f":${letter}${end_row}"
    )


def write_os_summary_sheet(wb, vm_df: pd.DataFrame) -> None:
    if vm_df is None or vm_df.empty:
        return

    ws = wb.create_sheet(title=OS_SUMMARY_SHEET)

    vd_end = columns.vd_data_end(len(vm_df))
    vd_os_range = _vd_abs_range("os_detected", vd_end)
    vd_cf_range = _vd_abs_range("oci_compatible", vd_end)

    ws.column_dimensions["A"].width = 37.1
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14

    # Row 1: title
    title_cell = ws.cell(row=1, column=1, value="OS Summary")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.row_dimensions[1].height = TITLE_ROW_HEIGHT

    cursor = 3  # row 2 intentionally blank

    # ── Section 1: OCI Compatibility Overview ────────────────────────────────
    lbl = ws.cell(row=cursor, column=1, value="OCI Compatibility Overview")
    lbl.font = Font(bold=True)
    lbl.fill = SECTION_LABEL_FILL
    lbl.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=3)
    ws.row_dimensions[cursor].height = DEFAULT_ROW_HEIGHT
    cursor += 1

    for col_idx, hdr in enumerate(["Status", "VM Count", "VM Percent"], start=1):
        cell = ws.cell(row=cursor, column=col_idx, value=hdr)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cursor].height = HEADER_ROW_HEIGHT
    cursor += 1

    compat_data_start = cursor
    compat_total_row = compat_data_start + 4  # yes/maybe/no/unknown + total

    for status in ("yes", "maybe", "no", "unknown"):
        ws.cell(row=cursor, column=1, value=status)
        ws.cell(row=cursor, column=2, value=f'=COUNTIF({vd_cf_range},"{status}")')
        pct_cell = ws.cell(
            row=cursor, column=3,
            value=f"=IF(B${compat_total_row}=0,0,B{cursor}/B${compat_total_row})",
        )
        pct_cell.number_format = "0.0%"
        ws.row_dimensions[cursor].height = DEFAULT_ROW_HEIGHT
        cursor += 1
    compat_data_end = cursor - 1

    total_cell = ws.cell(row=cursor, column=1, value="Total")
    total_cell.font = Font(bold=True)
    count_cell = ws.cell(row=cursor, column=2, value=f"=SUM(B{compat_data_start}:B{compat_data_end})")
    count_cell.font = Font(bold=True)
    ws.row_dimensions[cursor].height = DEFAULT_ROW_HEIGHT
    cursor += 2

    add_compat_rules(ws, f"A{compat_data_start}:A{compat_data_end}")

    # ── Section 2: OS Breakdown ───────────────────────────────────────────────
    lbl2 = ws.cell(row=cursor, column=1, value="OS Breakdown")
    lbl2.font = Font(bold=True)
    lbl2.fill = SECTION_LABEL_FILL
    lbl2.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=3)
    ws.row_dimensions[cursor].height = DEFAULT_ROW_HEIGHT
    cursor += 1

    for col_idx, hdr in enumerate(["OS Detected", "OCI Compatible", "VM Count"], start=1):
        cell = ws.cell(row=cursor, column=col_idx, value=hdr)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cursor].height = HEADER_ROW_HEIGHT
    cursor += 1

    # Unique OS list with counts, sorted by count desc then name asc.
    os_data: Dict[str, Tuple[str, int]] = {}
    for _, row in vm_df.iterrows():
        os_str, oci_compat, _ = detect_os(row)
        if os_str not in os_data:
            os_data[os_str] = (oci_compat, 0)
        os_data[os_str] = (os_data[os_str][0], os_data[os_str][1] + 1)

    sorted_os = sorted(os_data.items(), key=lambda x: (-x[1][1], x[0].lower()))

    os_data_start = cursor
    for os_str, (oci_compat, _) in sorted_os:
        display_os = os_str if os_str else "(No OS data)"
        ws.cell(row=cursor, column=1, value=display_os)
        ws.cell(row=cursor, column=2, value=oci_compat)
        if os_str:
            ws.cell(row=cursor, column=3, value=f"=COUNTIF({vd_os_range},A{cursor})")
        else:
            ws.cell(row=cursor, column=3, value=f'=COUNTIF({vd_os_range},"")')
        ws.row_dimensions[cursor].height = DEFAULT_ROW_HEIGHT
        cursor += 1
    os_data_end = cursor - 1

    total_os = ws.cell(row=cursor, column=1, value="Total")
    total_os.font = Font(bold=True)
    total_os_count = ws.cell(row=cursor, column=3, value=f"=SUM(C{os_data_start}:C{os_data_end})")
    total_os_count.font = Font(bold=True)
    ws.row_dimensions[cursor].height = DEFAULT_ROW_HEIGHT

    add_compat_rules(ws, f"B{os_data_start}:B{os_data_end}")
