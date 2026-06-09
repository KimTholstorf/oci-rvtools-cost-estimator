# Copyright (c) 2026 Kim Tholstorf
# https://github.com/KimTholstorf/oci-rvtools-cost-estimator
# MIT License — see LICENSE file for details

"""
OCI Monthly Cost Calculator (direct RVTools input)

Reads one or more raw RVTools Excel exports, reproduces the aggregation logic
used by rvtools_summarizer.py for powered-on VM CPU/RAM and provisioned/used
disk capacity, then retrieves current OCI list prices from the CETools API and
produces a cost breakdown workbook.
"""

from __future__ import annotations

import argparse
import json
import math
from datetime import datetime
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple
from urllib import error as urlerror
from urllib import parse as urlparse
from urllib import request as urlrequest

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule


# =========================
# Version
# =========================

VERSION = "1.2.0"


# =========================
# CLI Defaults
# =========================

DEFAULT_INPUT = ""  # unused placeholder
DEFAULT_OUTPUT = "oci_cost_summary.xlsx"
DEFAULT_CURRENCY = "USD"
DEFAULT_HOURS = 730
DEFAULT_OCPU_PART = "B97384"
DEFAULT_MEMORY_PART = "B97385"
DEFAULT_STORAGE_PART = "B91961"
DEFAULT_VPU_PART = "B91962"
DEFAULT_VPU = 10.0


# =========================
# API
# =========================

API_BASE = "https://apexapps.oracle.com/pls/apex/cetools/api/v1/products/"


# =========================
# vInfo Column Mapping
# =========================

CANON_COLS_VINFO = [
    "Powerstate",
    "VM",
    "Datacenter",
    "Cluster",
    "CPUs",
    "Memory",
    "Video Ram KiB",
    "Provisioned MiB",
    "In Use MiB",
    "Total disk capacity MiB",
    "OS according to the VMware Tools",
    "OS according to the configuration file",
]

ALIASES_VINFO = {
    "vInfoVMName": "VM",
    "vInfoPowerstate": "Powerstate",
    "vInfoDataCenter": "Datacenter",
    "vInfoCluster": "Cluster",
    "vInfoHost": "Host",
    "vInfoCPUs": "CPUs",
    "vInfoMemory": "Memory",
    "vInfoVideoRamKiB": "Video Ram KiB",
    "vInfoProvisioned": "Provisioned MiB",
    "vInfoInUse": "In Use MiB",
    "vInfoTotalDiskCapacityMiB": "Total disk capacity MiB",
    "vInfoPrimaryIPAddress": "Primary IP Address",
    "vInfoPrimaryIP": "Primary IP Address",
    "vInfoOS": "OS according to the VMware Tools",
    "vInfoOSTools": "OS according to the VMware Tools",
    "vInfoOSConfig": "OS according to the configuration file",
}

TOKEN_MAP_VINFO = {
    "powerstate": "Powerstate",
    "vm": "VM",
    "vmname": "VM",
    "datacenter": "Datacenter",
    "datacentre": "Datacenter",
    "cluster": "Cluster",
    "cpus": "CPUs",
    "memory": "Memory",
    "videorammikb": "Video Ram KiB",
    "videorammikib": "Video Ram KiB",
    "provisioned": "Provisioned MiB",
    "provisionedmib": "Provisioned MiB",
    "inuse": "In Use MiB",
    "inusemib": "In Use MiB",
    "totaldiskcapacitymib": "Total disk capacity MiB",
    "osaccordingtovmwaretools": "OS according to the VMware Tools",
    "osaccordingtotheconfigurationfile": "OS according to the configuration file",
}

NUMERIC_PREFERRED_VINFO = {
    "CPUs",
    "Memory",
    "Video Ram KiB",
    "Provisioned MiB",
    "In Use MiB",
    "Total disk capacity MiB",
}

INVALID_CLUSTER_VALUES = {"", "none", "nan", "unknown"}


# =========================
# Unit Conversion
# =========================

MIB_TO_GB = 1024.0 / 953_674.0  # RVTools MiB -> TiB -> GB conversion factor


# =========================
# Excel Layout
# =========================

TABLE_COLUMN_COUNT = 7
EXCEL_HEADERS = [
    "Description",
    "Part Number",
    "Part Qty",
    "Instance Qty",
    "Usage Qty",
    "Unit Price ({currency})",
    "Monthly Cost ({currency})",
]
COLUMN_WIDTHS = {
    "A": 36,
    "B": 16,
    "C": 16,
    "D": 16,
    "E": 16,
    "F": 16,
    "G": 16,
}
TITLE_ROW_HEIGHT = 40
HEADER_ROW_HEIGHT = 40
DEFAULT_ROW_HEIGHT = 20
DISCLAIMER_ROW_HEIGHT = 80

VM_DETAIL_HEADERS = [
    "VM Name",
    "Powerstate",
    "OS Detected",
    "OCI Compatible",
    "OCPU",
    "RAM (GB)",
    "Disk Provisioned (GB)",
    "Disk Used (GB)",
    "OCPU Cost ({currency})",
    "RAM Cost ({currency})",
    "Disk Prov Cost ({currency})",
    "Disk Used Cost ({currency})",
    "Total Monthly - Prov ({currency})",
    "Total Monthly - Used ({currency})",
    "Total Yearly - Prov ({currency})",
    "Total Yearly - Used ({currency})",
    "Note",
]

VM_DETAIL_COL_WIDTHS = {
    "A": 34, "B": 14, "C": 37.1, "D": 16,
    "E": 10, "F": 12, "G": 22, "H": 18,
    "I": 22, "J": 22, "K": 24, "L": 24,
    "M": 28, "N": 26, "O": 28, "P": 26,
    "Q": 16,
}

VM_DETAIL_NOTE = (
    "ⓘ  The aggregate totals on the total_disk and used_disk sheets are calculated by "
    "summing the per-VM values in this sheet. Editing a value here will automatically "
    "update the monthly cost totals."
)


# =========================
# Excel Styles
# =========================

TITLE_FONT = Font(bold=True, size=14)
HEADER_FILL = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
SECTION_LABEL_FILL = PatternFill(start_color="FFB8CCE4", end_color="FFB8CCE4", fill_type="solid")
DISCLAIMER_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

SECTION_LABELS = {
    "total_disk": "Total Provisioned Disk",
    "used_disk":  "Total Used Disk",
}

# OCI compatibility conditional formatting fills — applied as Excel rules so
# the colour updates automatically if a cell value is edited.
_CF_FILL_YES     = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")  # soft green
_CF_FILL_MAYBE   = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")  # soft amber
_CF_FILL_NO      = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")  # soft red

# Ordered rules for OS → OCI compatibility classification.
# Each entry: (lowercase_substring, result, note)
# First match wins; "unknown" is returned when nothing matches.
OS_OCI_COMPAT_RULES: List[Tuple[str, str, str]] = [
    # ── Definite "no" ──────────────────────────────────────────────────────────
    ("vmware esx",                      "no",    ""),
    ("esxi",                            "no",    ""),
    ("mac os",                          "no",    ""),
    ("macos",                           "no",    ""),
    ("solaris",                         "no",    ""),
    ("netware",                         "no",    ""),
    ("ms-dos",                          "no",    ""),
    ("windows xp",                      "no",    ""),
    ("windows vista",                   "no",    ""),
    ("windows 7 ",                      "no",    ""),
    ("windows 7)",                      "no",    ""),
    ("windows 8 ",                      "no",    ""),
    ("windows 8)",                      "no",    ""),
    ("(32-bit)",                        "no",    "32-bit OS not supported on OCI"),
    # ── Windows Desktop → "maybe" (OCI Secure Desktop only) ───────────────────
    ("windows 10",                      "maybe", "Only supported on OCI Secure Desktop"),
    ("windows 11",                      "maybe", "Only supported on OCI Secure Desktop"),
    # ── Windows Server specific versions → "yes" ──────────────────────────────
    ("windows server 2025",             "yes",   ""),
    ("windows server 2022",             "yes",   ""),
    ("windows server 2019",             "yes",   ""),
    ("windows server 2016",             "yes",   ""),
    # ── Windows Server old/unsupported versions → "no" ────────────────────────
    ("windows server 2012",             "no",    ""),
    ("windows server 2008",             "no",    ""),
    ("windows server 2003",             "no",    ""),
    ("windows server 2000",             "no",    ""),
    ("windows nt",                      "no",    ""),
    # ── Windows Server generic (unknown/future version) → "maybe" ─────────────
    ("windows server",                  "maybe", ""),
    # ── Oracle Linux / Autonomous → "yes" ─────────────────────────────────────
    ("oracle autonomous linux",         "yes",   ""),
    ("oracle linux",                    "yes",   ""),
    # ── Ubuntu: specific supported LTS → "yes"; generic → "maybe" ─────────────
    ("ubuntu linux 20",                 "yes",   ""),
    ("ubuntu linux 22",                 "yes",   ""),
    ("ubuntu linux 24",                 "yes",   ""),
    ("ubuntu 20",                       "yes",   ""),
    ("ubuntu 22",                       "yes",   ""),
    ("ubuntu 24",                       "yes",   ""),
    ("ubuntu",                          "maybe", ""),
    # ── RHEL: specific versions → "yes"; generic/old → "maybe" ───────────────
    ("red hat enterprise linux 6",      "yes",   ""),
    ("red hat enterprise linux 7",      "yes",   ""),
    ("red hat enterprise linux 8",      "yes",   ""),
    ("red hat enterprise linux 9",      "yes",   ""),
    ("red hat enterprise linux 2",      "no",    ""),
    ("red hat enterprise linux 3",      "no",    ""),
    ("red hat enterprise linux",        "maybe", ""),
    # ── CentOS: modern specific → "yes"; old/generic → "maybe" ───────────────
    ("centos stream",                   "yes",   ""),
    ("centos 7",                        "yes",   ""),
    ("centos 6",                        "yes",   ""),
    ("centos",                          "maybe", ""),
    # ── SUSE: specific versions → "yes"; generic → "maybe" ───────────────────
    ("suse linux enterprise 11",        "yes",   ""),
    ("suse linux enterprise 12",        "yes",   ""),
    ("suse linux enterprise 15",        "yes",   ""),
    ("suse linux enterprise",           "maybe", ""),
    ("opensuse",                        "yes",   ""),
    # ── Debian: versions 8+ → "yes"; generic → "maybe" ───────────────────────
    ("debian gnu/linux 8",              "yes",   ""),
    ("debian gnu/linux 9",              "yes",   ""),
    ("debian gnu/linux 10",             "yes",   ""),
    ("debian gnu/linux 11",             "yes",   ""),
    ("debian gnu/linux 12",             "yes",   ""),
    ("debian",                          "maybe", ""),
    # ── FreeBSD ───────────────────────────────────────────────────────────────
    ("freebsd",                         "maybe", ""),
    # ── Catch-all Linux / Other ───────────────────────────────────────────────
    ("other linux",                     "maybe", ""),
    ("linux",                           "maybe", ""),
    ("other (64-bit)",                  "maybe", ""),
    ("other",                           "maybe", ""),
]


# =========================
# Logging helpers
# =========================

def info(msg: str) -> None:
    print(f"[INFO] {msg}")


def warn(msg: str) -> None:
    print(f"[WARN] {msg}")


# =========================
# OS detection helper
# =========================

def detect_os(row_data: "pd.Series") -> Tuple[str, str, str]:
    """Return (os_detected, oci_compatible, compat_note) for a vInfo row.

    os_detected   – raw OS string from RVTools (VMware Tools column, falling
                    back to configuration file column when the first is empty).
    oci_compatible – "yes" | "maybe" | "no" | "unknown"
    compat_note   – short human-readable note, e.g. "Only supported on OCI Secure Desktop"
    """
    os_str = ""
    for col in ("OS according to the VMware Tools", "OS according to the configuration file"):
        raw = row_data.get(col, "")
        if pd.notna(raw) and str(raw).strip():
            os_str = str(raw).strip()
            break

    if not os_str:
        return "", "unknown", ""

    lower = os_str.lower()
    for pattern, result, note in OS_OCI_COMPAT_RULES:
        if pattern in lower:
            return os_str, result, note

    return os_str, "unknown", ""


# =========================
# RVTools helpers
# =========================

def _sheet_token(name: str) -> str:
    return "".join(ch for ch in (name or "").strip().lower() if ch.isalnum())


def _to_token(value: str) -> str:
    return "".join(ch for ch in value.strip().lower() if ch.isalnum() or ch == "#")


def collapse_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Coalesce duplicate column names (sum numeric, else first non-empty)."""
    if df is None or df.empty:
        return df

    cols = list(df.columns)
    duplicates = {c for c in cols if cols.count(c) > 1}
    for col in duplicates:
        numeric = col in NUMERIC_PREFERRED_VINFO
        parts = [df.iloc[:, idx] for idx, name in enumerate(cols) if name == col]
        stacked = pd.concat(parts, axis=1)
        if numeric:
            coalesced = pd.to_numeric(stacked, errors="coerce").fillna(0).sum(axis=1)
        else:
            coalesced = stacked.apply(
                lambda row: next((v for v in row if pd.notna(v) and str(v).strip()), ""),
                axis=1,
            )
        first_idx = next(i for i, name in enumerate(cols) if name == col)
        mask = [True] * len(cols)
        seen = 0
        for i, name in enumerate(cols):
            if name == col:
                seen += 1
                if seen > 1:
                    mask[i] = False
        df = df.loc[:, mask]
        cols = list(df.columns)
        df.iloc[:, first_idx] = coalesced
        info(f"vInfo: coalesced duplicate column '{col}' ({len(parts)} -> 1)")
    return df


def canonicalize_vinfo(df: pd.DataFrame) -> pd.DataFrame:
    if not isinstance(df, pd.DataFrame):
        return df

    # Aliases
    aliases = {old: new for (old, new) in ALIASES_VINFO.items() if old in df.columns}
    if aliases:
        df = df.rename(columns=aliases)
        for old, new in aliases.items():
            info(f"vInfo: renaming column '{old}' -> '{new}' (alias)")

    # Token-based normalization
    renames: Dict[str, str] = {}
    for col in df.columns:
        token = _to_token(str(col))
        target = TOKEN_MAP_VINFO.get(token)
        if target and target != col:
            renames[col] = target
    if renames:
        df = df.rename(columns=renames)
        for old, new in renames.items():
            info(f"vInfo: renaming column '{old}' -> '{new}' (token)")

    df = collapse_duplicate_columns(df)

    # Backfill required columns
    for column in CANON_COLS_VINFO:
        if column not in df.columns:
            warn(f"vInfo: missing column '{column}', backfilling default")
            df[column] = 0 if column in NUMERIC_PREFERRED_VINFO else ""

    return df


_OS_COLS = ["OS according to the VMware Tools", "OS according to the configuration file"]


def load_vinfo_dataframe(filepath: Path) -> Optional[pd.DataFrame]:
    try:
        xl = pd.read_excel(filepath, sheet_name=None, engine="openpyxl")
    except Exception as exc:
        warn(f"{filepath.name}: failed to load workbook ({exc})")
        return None

    vinfo_df = None
    vcpu_df = None
    for sheet_name, df in xl.items():
        token = _sheet_token(sheet_name)
        if token.endswith("vinfo") or "vinfo" in token:
            vinfo_df = df
        elif token == "vcpu":
            vcpu_df = df

    if vinfo_df is None:
        warn(f"{filepath.name}: missing vInfo sheet, skipping")
        return None

    vinfo_df = canonicalize_vinfo(vinfo_df)

    # If vInfo is missing OS columns, try to pull them from vCPU (common in trimmed exports)
    missing_os = [c for c in _OS_COLS if c not in vinfo_df.columns or vinfo_df[c].astype(str).str.strip().eq("").all()]
    if missing_os and vcpu_df is not None and "VM" in vcpu_df.columns:
        available = [c for c in _OS_COLS if c in vcpu_df.columns]
        if available:
            os_lookup = (
                vcpu_df[["VM"] + available]
                .drop_duplicates(subset="VM")
                .rename(columns={c: c for c in available})
            )
            vinfo_df = vinfo_df.merge(os_lookup, on="VM", how="left", suffixes=("", "_vcpu"))
            # Prefer the freshly merged column; drop the _vcpu suffixed duplicate if any
            for col in available:
                vcpu_col = f"{col}_vcpu"
                if vcpu_col in vinfo_df.columns:
                    vinfo_df[col] = vinfo_df[col].where(
                        vinfo_df[col].astype(str).str.strip().ne(""), vinfo_df[vcpu_col]
                    )
                    vinfo_df.drop(columns=[vcpu_col], inplace=True)
            info(f"vInfo: pulled OS columns from vCPU sheet ({', '.join(available)})")

    return vinfo_df


# =========================
# Aggregation from vInfo
# =========================

def _valid_cluster(value: object) -> bool:
    if pd.isna(value):
        return False
    s = str(value).strip()
    if not s:
        return False
    return s.lower() not in INVALID_CLUSTER_VALUES


def _prepare_vinfo_df(df: pd.DataFrame, verbose: bool = True) -> pd.DataFrame:
    """Remove vCLS housekeeping VMs and apply cluster filter when cluster data exists."""
    df = df.copy()
    df = df[~df["VM"].astype(str).str.startswith("vCLS", na=False)]
    if df["Cluster"].apply(_valid_cluster).any():
        df = df[df["Cluster"].apply(_valid_cluster)]
    elif verbose:
        info("No valid cluster data detected — cluster filter skipped, all VMs included")
    return df


@dataclass
class AggregatedUsage:
    total_vcpu: float = 0.0
    ram_gb: float = 0.0
    disk_total_gb: float = 0.0
    disk_used_gb: float = 0.0
    powered_on_vms: int = 0
    powered_off_vms: int = 0
    source_files: List[str] = field(default_factory=list)
    vm_dataframe: Optional[pd.DataFrame] = field(default=None)


def aggregate_vinfo(df: pd.DataFrame, include_vms_off: bool, include_disks_off: bool) -> Tuple[float, float, float, float, int, int]:
    if df is None or df.empty:
        return 0.0, 0.0, 0.0, 0.0, 0, 0

    df = _prepare_vinfo_df(df)
    if df.empty:
        return 0.0, 0.0, 0.0, 0.0, 0, 0

    powered_mask = df["Powerstate"].astype(str).str.lower() == "poweredon"
    powered_on_count = int(powered_mask.sum())
    total_count = int(len(df))
    powered_off_count = total_count - powered_on_count

    vcpu_frame = df if include_vms_off else df[powered_mask]
    total_vcpu = float(pd.to_numeric(vcpu_frame["CPUs"], errors="coerce").fillna(0).sum())
    ram_gb = float(pd.to_numeric(vcpu_frame["Memory"], errors="coerce").fillna(0).sum()) / 1024.0

    disk_frame = df if include_disks_off else df[powered_mask]
    total_mib = pd.to_numeric(disk_frame["Total disk capacity MiB"], errors="coerce").fillna(0)
    prov_mib = pd.to_numeric(disk_frame["Provisioned MiB"], errors="coerce").fillna(0)
    effective_prov = total_mib.where(total_mib > 0, prov_mib)
    used_mib = pd.to_numeric(disk_frame["In Use MiB"], errors="coerce").fillna(0)

    disk_total_gb = float(effective_prov.sum() * MIB_TO_GB)
    disk_used_gb = float(used_mib.sum() * MIB_TO_GB)

    return total_vcpu, ram_gb, disk_total_gb, disk_used_gb, powered_on_count, powered_off_count


def collect_rvtools_files(paths: Sequence[str]) -> List[Path]:
    files: List[Path] = []
    seen: set[Path] = set()
    skipped_temp: set[str] = set()

    def add_candidate(candidate: Path) -> None:
        resolved = candidate.resolve()
        if resolved in seen:
            return
        seen.add(resolved)
        files.append(candidate)

    for entry in paths:
        p = Path(entry).expanduser()
        if p.is_dir():
            for candidate in sorted(p.glob("*.xlsx")):
                if candidate.name.startswith("~$"):
                    skipped_temp.add(candidate.name)
                    continue
                add_candidate(candidate)
        elif p.is_file():
            if p.name.startswith("~$"):
                skipped_temp.add(p.name)
                continue
            add_candidate(p)
        else:
            warn(f"Input path not found: {p}")

    if skipped_temp:
        info(f"Skipped temporary workbook(s): {', '.join(sorted(skipped_temp))}")
    return files


def apply_vm_filter(
    df: pd.DataFrame,
    datacenters: Optional[List[str]],
    clusters: Optional[List[str]],
) -> pd.DataFrame:
    """Filter vInfo rows by Datacenter and/or Cluster names (case-insensitive, AND logic between flags)."""
    if not datacenters and not clusters:
        return df

    mask = pd.Series(True, index=df.index)

    if datacenters:
        dc_lower = [d.lower() for d in datacenters]
        df_dc = df["Datacenter"].astype(str).str.strip().str.lower()
        matched = set(df_dc[df_dc.isin(dc_lower)].unique())
        for d in dc_lower:
            if d not in matched:
                warn(f"No VMs found for Datacenter '{d}' — check spelling and case")
        mask &= df_dc.isin(dc_lower)

    if clusters:
        cl_lower = [c.lower() for c in clusters]
        df_cl = df["Cluster"].astype(str).str.strip().str.lower()
        matched = set(df_cl[df_cl.isin(cl_lower)].unique())
        for c in cl_lower:
            if c not in matched:
                warn(f"No VMs found for Cluster '{c}' — check spelling and case")
        mask &= df_cl.isin(cl_lower)

    return df[mask]


def list_datacenters_and_clusters(files: Sequence[Path]) -> None:
    """Print all unique Datacenter and Cluster names found across the given RVTools files."""
    topology: Dict[str, set] = {}
    for file in files:
        df = load_vinfo_dataframe(file)
        if df is None:
            continue
        df = df[df["Cluster"].apply(_valid_cluster)]
        for _, row in df[["Datacenter", "Cluster"]].drop_duplicates().iterrows():
            dc = str(row["Datacenter"]).strip() or "(unknown)"
            cl = str(row["Cluster"]).strip()
            topology.setdefault(dc, set()).add(cl)
    if not topology:
        print("[WARN] No Datacenter/Cluster data found in the provided files.")
        return
    for dc in sorted(topology):
        print(f"Datacenter: {dc}")
        for cl in sorted(topology[dc]):
            print(f"  Cluster: {cl}")


def aggregate_from_rvtools(
    files: Sequence[Path],
    include_vms_off: bool,
    include_disks_off: bool,
    datacenters: Optional[List[str]] = None,
    clusters: Optional[List[str]] = None,
) -> AggregatedUsage:
    usage = AggregatedUsage()
    vm_frames: List[pd.DataFrame] = []
    for file in files:
        info(f"Processing RVTools export: {file}")
        df = load_vinfo_dataframe(file)
        if df is None:
            continue
        df = apply_vm_filter(df, datacenters, clusters)
        if df.empty:
            warn(f"{file.name}: no VMs remain after applying datacenter/cluster filter, skipping")
            continue
        detail_df = _prepare_vinfo_df(df, verbose=False)
        if not detail_df.empty:
            vm_frames.append(detail_df)
        total_vcpu, ram_gb, disk_total_gb, disk_used_gb, powered_on, powered_off = aggregate_vinfo(
            df, include_vms_off, include_disks_off
        )
        usage.total_vcpu += total_vcpu
        usage.ram_gb += ram_gb
        usage.disk_total_gb += disk_total_gb
        usage.disk_used_gb += disk_used_gb
        usage.powered_on_vms += powered_on
        usage.powered_off_vms += powered_off
        usage.source_files.append(str(file))

    if vm_frames:
        usage.vm_dataframe = pd.concat(vm_frames, ignore_index=True)

    return usage


# =========================
# Pricing client
# =========================

@dataclass
class LineItem:
    description: str
    part_number: str
    category: str
    raw_base_quantity: float
    usage_quantity: float
    unit_price: float


@dataclass
class PriceRecord:
    part_number: str
    display_name: str
    unit_price: float


class PricingClient:
    def __init__(self, currency: str) -> None:
        self.currency = currency.upper()
        self._cache: Dict[str, PriceRecord] = {}

    def get_price(self, part_number: str) -> PriceRecord:
        part_number = part_number.strip()
        if part_number in self._cache:
            return self._cache[part_number]

        params = {"partNumber": part_number, "currencyCode": self.currency}
        url = f"{API_BASE}?{urlparse.urlencode(params)}"
        try:
            with urlrequest.urlopen(url) as response:
                payload = response.read().decode("utf-8")
        except urlerror.URLError as exc:
            raise RuntimeError(f"Failed to reach OCI price API ({url}): {exc}") from exc

        try:
            data = json.loads(payload)
        except json.JSONDecodeError as exc:
            raise RuntimeError(f"Invalid JSON from OCI price API for part {part_number}") from exc

        items = data.get("items")
        if not isinstance(items, list) or not items:
            raise RuntimeError(f"No pricing data for part {part_number} ({self.currency})")

        price: Optional[float] = None
        display_name: Optional[str] = None
        for item in items:
            if not isinstance(item, dict):
                continue
            if item.get("partNumber") and item["partNumber"] != part_number:
                continue
            price = self._extract_price(item, self.currency)
            display_name = item.get("displayName") or display_name
            if price is not None:
                break

        if price is None:
            for item in items:
                if isinstance(item, dict):
                    price = PricingClient._extract_price(item, self.currency)
                    display_name = item.get("displayName") or display_name
                    if price is not None:
                        break

        if price is None:
            raise RuntimeError(f"Could not determine unit price for part {part_number}: {payload[:500]}...")

        if not display_name:
            display_name = part_number

        record = PriceRecord(part_number=part_number, display_name=str(display_name), unit_price=float(price))
        self._cache[part_number] = record
        return record

    @staticmethod
    def _extract_price(item: Dict[str, object], currency: str) -> Optional[float]:
        currency = currency.upper()
        candidate_keys = [
            "price",
            "unitPrice",
            "unit_price",
            "unit_price_value",
            "unit_price_included",
            "netUnitPrice",
            "list_price",
            "usdPrice",
            "amount",
        ]
        for key in candidate_keys:
            if key in item:
                try:
                    return float(item[key])  # type: ignore[arg-type]
                except (TypeError, ValueError):
                    continue

        if "prices" in item and isinstance(item["prices"], list):
            for entry in item["prices"]:
                if isinstance(entry, dict):
                    maybe = PricingClient._extract_price(entry, currency)
                    if maybe is not None:
                        return maybe

        locs = item.get("currencyCodeLocalizations")
        if isinstance(locs, list):
            for loc in locs:
                if not isinstance(loc, dict):
                    continue
                loc_currency = loc.get("currencyCode")
                if loc_currency and loc_currency.upper() != currency:
                    continue
                prices = loc.get("prices")
                if isinstance(prices, list):
                    for entry in prices:
                        if not isinstance(entry, dict):
                            continue
                        if "value" in entry:
                            try:
                                return float(entry["value"])
                            except (TypeError, ValueError):
                                pass
                        maybe = PricingClient._extract_price(entry, currency)
                        if maybe is not None:
                            return maybe
        return None


def build_line_items(
    usage: AggregatedUsage,
    hours: float,
    vpu_value: float,
    pricing: PricingClient,
    parts: Dict[str, str],
) -> Dict[str, List[LineItem]]:
    raw_total_ocpus = usage.total_vcpu / 2.0
    raw_memory_gb = usage.ram_gb
    raw_total_disk_gb = usage.disk_total_gb
    raw_used_disk_gb = usage.disk_used_gb

    total_ocpus = math.ceil(raw_total_ocpus)
    memory_gb = math.ceil(raw_memory_gb)
    total_disk_gb = math.ceil(raw_total_disk_gb)
    used_disk_gb = math.ceil(raw_used_disk_gb)

    ocpu_hours = total_ocpus * hours
    memory_hours = memory_gb * hours

    info(f"Total OCPUs: {total_ocpus:.2f}, OCPU hours: {ocpu_hours:.2f}")
    info(f"Memory GB: {memory_gb:.2f}, GB hours: {memory_hours:.2f}")
    info(f"Total disk GB: {total_disk_gb:.2f}")
    info(f"Used disk GB: {used_disk_gb:.2f}")
    info(f"Hours per month: {hours}")
    info(f"VPU per GB: {vpu_value}")

    price_records = {
        key: pricing.get_price(parts[key])
        for key in ("ocpu", "memory", "storage", "vpu")
    }

    def make_lines(disk_gb: float, raw_disk_gb: float) -> List[LineItem]:
        raw_vpu_quantity = raw_disk_gb * vpu_value
        return [
            LineItem(
                description=price_records["ocpu"].display_name,
                part_number=price_records["ocpu"].part_number,
                category="ocpu",
                raw_base_quantity=raw_total_ocpus,
                usage_quantity=hours,
                unit_price=price_records["ocpu"].unit_price,
            ),
            LineItem(
                description=price_records["memory"].display_name,
                part_number=price_records["memory"].part_number,
                category="memory",
                raw_base_quantity=raw_memory_gb,
                usage_quantity=hours,
                unit_price=price_records["memory"].unit_price,
            ),
            LineItem(
                description=price_records["storage"].display_name,
                part_number=price_records["storage"].part_number,
                category="storage",
                raw_base_quantity=raw_disk_gb,
                usage_quantity=1.0,
                unit_price=price_records["storage"].unit_price,
            ),
            LineItem(
                description=price_records["vpu"].display_name,
                part_number=price_records["vpu"].part_number,
                category="vpu",
                raw_base_quantity=raw_vpu_quantity,
                usage_quantity=1.0,
                unit_price=price_records["vpu"].unit_price,
            ),
        ]

    return {
        "total_disk": make_lines(total_disk_gb, raw_total_disk_gb),
        "used_disk": make_lines(used_disk_gb, raw_used_disk_gb),
    }


# =========================
# Excel output
# =========================

def build_metadata_rows(metadata: Dict[str, object]) -> List[Tuple[str, str, str]]:
    filter_parts = []
    if metadata.get("filter_datacenters"):
        filter_parts.append(f"Datacenter: {', '.join(metadata['filter_datacenters'])}")
    if metadata.get("filter_clusters"):
        filter_parts.append(f"Cluster: {', '.join(metadata['filter_clusters'])}")
    filter_value = " | ".join(filter_parts) if filter_parts else "None"

    return [
        ("Source Files", metadata["source_files"], ""),
        ("Filters", filter_value, ""),
        ("Hours per Month", metadata["hours_per_month"], ""),
        ("Currency", metadata["currency"], ""),
        ("VPU", metadata["vpu"], ""),
        ("Powered On VMs", str(metadata["powered_on_vms"]), "(included)"),
        (
            "Powered Off VMs",
            str(metadata["powered_off_vms"]),
            "(included)" if metadata["powered_off_included"] else "(excluded)",
        ),
        (
            "Powered Off Disks",
            "",
            "(included)" if metadata["powered_off_disks_included"] else "(excluded)",
        ),
    ]


def apply_column_widths(ws) -> None:
    for column, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width


def write_metadata(ws, rows: List[Tuple[str, str, str]], start_row: int = 2) -> int:
    row = start_row
    for title, value, status in rows:
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=2, value=value)
        if status:
            ws.cell(row=row, column=3, value=status)
        row += 1
    return row - 1


def format_row(ws, row: int, *, bold: bool = False, vertical: str = "bottom") -> None:
    for col_idx in range(1, TABLE_COLUMN_COUNT + 1):
        cell = ws.cell(row=row, column=col_idx)
        if bold:
            cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical=vertical)


def format_metadata_block(ws, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row + 1):
        format_row(ws, row)
    format_row(ws, start_row, bold=True)


def format_table_header(ws, header_row: int, headers: List[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = HEADER_FILL


def format_data_rows(ws, start_row: int, end_row: int) -> None:
    for row in range(start_row, end_row):
        format_row(ws, row)


def format_total_row(ws, total_row: int) -> None:
    format_row(ws, total_row, bold=True)


def apply_row_heights(
    ws,
    metadata_end_row: int,
    blank_row_index: int,
    header_row: int,
    data_start: int,
    total_row: int,
    disclaimer_row: int,
) -> None:
    ws.row_dimensions[1].height = TITLE_ROW_HEIGHT
    for row in range(2, metadata_end_row + 1):
        ws.row_dimensions[row].height = DEFAULT_ROW_HEIGHT
    ws.row_dimensions[blank_row_index].height = DEFAULT_ROW_HEIGHT
    ws.row_dimensions[header_row].height = HEADER_ROW_HEIGHT
    for row in range(data_start, total_row + 1):
        ws.row_dimensions[row].height = DEFAULT_ROW_HEIGHT
    quote_row = disclaimer_row - 2
    if quote_row >= 2:
        ws.row_dimensions[quote_row].height = DEFAULT_ROW_HEIGHT
    ws.row_dimensions[disclaimer_row].height = DISCLAIMER_ROW_HEIGHT


def append_advisory(ws, total_row: int) -> Tuple[int, int]:
    quote_row = total_row + 3
    quote_cell = ws.cell(row=quote_row, column=1, value="Quote is for investment proposal only.")
    quote_cell.font = Font(bold=True, color="00FF0000")
    ws.merge_cells(start_row=quote_row, start_column=1, end_row=quote_row, end_column=TABLE_COLUMN_COUNT)

    disclaimer_row = quote_row + 2
    disclaimer_text = (
        "Disclaimer:  This sample quote is provided solely for evaluation purposes and is intended to further "
        "discussions between you and Oracle.  This sample quote is not eligible for acceptance by you and is "
        "not a binding contract between you and Oracle for the services specified.  If you would like to purchase "
        "the services specified in this sample quote, please request that Oracle issue you a formal quote (which "
        "may include an OMA or a CSA if you do not already have an appropriate agreement in place with Oracle) "
        "for your acceptance and execution.  Your formal quote will be effective only upon Oracle's acceptance "
        "of the formal quote (and the OMA or CSA, if required)."
    )
    disclaimer_cell = ws.cell(row=disclaimer_row, column=1, value=disclaimer_text)
    ws.merge_cells(start_row=disclaimer_row, start_column=1, end_row=disclaimer_row, end_column=TABLE_COLUMN_COUNT)
    disclaimer_cell.alignment = DISCLAIMER_ALIGNMENT
    ws.row_dimensions[disclaimer_row].height = DISCLAIMER_ROW_HEIGHT
    return quote_row, disclaimer_row


def write_vm_detail_sheet(
    wb,
    vm_df: pd.DataFrame,
    metadata: Dict[str, object],
    currency: str,
    include_vms_off: bool,
    include_disks_off: bool,
) -> None:
    """Append a 'VM Details' sheet with one row per VM and per-VM cost formulas."""
    if vm_df is None or vm_df.empty:
        return

    ws = wb.create_sheet(title="VM Details")

    # Compute cross-sheet row references into 'total_disk'
    meta_rows = build_metadata_rows(metadata)
    n_meta = len(meta_rows)
    hours_row      = 2 + next(i for i, (k, _, _) in enumerate(meta_rows) if k == "Hours per Month")
    vpu_meta_row   = 2 + next(i for i, (k, _, _) in enumerate(meta_rows) if k == "VPU")
    data_start_ref = 2 + n_meta + 3   # title(1) + metadata(n_meta rows starting at 2) + blank + section_label + header
    ocpu_row  = data_start_ref         # OCPU line item row in total_disk
    ram_row   = data_start_ref + 1     # Memory line item row
    stor_row  = data_start_ref + 2     # Storage line item row
    vpu_row   = data_start_ref + 3     # VPU line item row

    td = "Cost Summary"  # source sheet name
    ocpu_price = f"'{td}'!F${ocpu_row}"       # OCPU unit price
    ram_price  = f"'{td}'!F${ram_row}"        # RAM unit price
    stor_price = f"'{td}'!F${stor_row}"       # storage unit price per GB
    vpu_pgb    = f"'{td}'!B${vpu_meta_row}"   # VPU per GB (metadata value)
    vpu_price  = f"'{td}'!F${vpu_row}"        # VPU unit price

    # Disk cost per GB = storage_price + vpu_per_gb * vpu_price
    disk_cost_per_gb = f"({stor_price}+{vpu_pgb}*{vpu_price})"

    for col, width in VM_DETAIL_COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    n_cols = len(VM_DETAIL_HEADERS)

    # Row 1: column headers
    headers = [h.format(currency=currency) for h in VM_DETAIL_HEADERS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.fill = HEADER_FILL
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    # Row 2: rounding note — visible immediately without scrolling
    note_cell = ws.cell(row=2, column=1, value=VM_DETAIL_NOTE)
    note_cell.font = Font(italic=True, color="00888888")
    note_cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    ws.row_dimensions[2].height = 36

    # Sort by VM name (case-insensitive); data starts at row 3
    df = vm_df.sort_values("VM", key=lambda s: s.str.lower()).reset_index(drop=True)

    for idx, row_data in df.iterrows():
        r = int(idx) + 3  # Excel row (row 1=headers, row 2=note, row 3+ = data)

        powerstate     = str(row_data["Powerstate"])
        is_powered_off = powerstate.strip().lower() != "poweredon"

        cpus      = float(pd.to_numeric(row_data["CPUs"],                      errors="coerce") or 0)
        mem_mib   = float(pd.to_numeric(row_data["Memory"],                    errors="coerce") or 0)
        total_mib = float(pd.to_numeric(row_data["Total disk capacity MiB"],   errors="coerce") or 0)
        prov_mib  = float(pd.to_numeric(row_data["Provisioned MiB"],           errors="coerce") or 0)
        used_mib  = float(pd.to_numeric(row_data["In Use MiB"],                errors="coerce") or 0)

        eff_prov_mib = total_mib if total_mib > 0 else prov_mib

        ocpu_val      = math.ceil(cpus / 2.0)
        ram_gb_val    = math.ceil(mem_mib / 1024.0)
        disk_prov_val = math.ceil(eff_prov_mib * MIB_TO_GB)
        disk_used_val = math.ceil(used_mib * MIB_TO_GB)

        include_cpu_ram = not is_powered_off or include_vms_off
        include_disk    = not is_powered_off or include_disks_off

        # OS detection
        os_detected, oci_compat, compat_note = detect_os(row_data)

        # Columns A–H: descriptive values
        ws.cell(row=r, column=1,  value=str(row_data["VM"]))
        ws.cell(row=r, column=2,  value=powerstate)
        ws.cell(row=r, column=3,  value=os_detected)                # OS Detected
        ws.cell(row=r, column=4,  value=oci_compat)                 # OCI Compatible
        ws.cell(row=r, column=5,  value=ocpu_val)
        ws.cell(row=r, column=6,  value=ram_gb_val)
        ws.cell(row=r, column=7,  value=disk_prov_val)
        ws.cell(row=r, column=8,  value=disk_used_val)

        # Columns I–L: cost formulas (or 0 when excluded by power-state flags)
        if include_cpu_ram:
            ws.cell(row=r, column=9,  value=f"=E{r}*'{td}'!B${hours_row}*{ocpu_price}")
            ws.cell(row=r, column=10, value=f"=F{r}*'{td}'!B${hours_row}*{ram_price}")
        else:
            ws.cell(row=r, column=9,  value=0)
            ws.cell(row=r, column=10, value=0)

        if include_disk:
            ws.cell(row=r, column=11, value=f"=G{r}*{disk_cost_per_gb}")
            ws.cell(row=r, column=12, value=f"=H{r}*{disk_cost_per_gb}")
        else:
            ws.cell(row=r, column=11, value=0)
            ws.cell(row=r, column=12, value=0)

        # Columns M–P: totals
        ws.cell(row=r, column=13, value=f"=I{r}+J{r}+K{r}")   # Total monthly prov
        ws.cell(row=r, column=14, value=f"=I{r}+J{r}+L{r}")   # Total monthly used
        ws.cell(row=r, column=15, value=f"=M{r}*12")           # Total yearly prov
        ws.cell(row=r, column=16, value=f"=N{r}*12")           # Total yearly used

        # Column Q: note — combine powerstate and OS compat notes where applicable
        note_parts = []
        if is_powered_off:
            note_parts.append("Powered Off")
        if compat_note:
            note_parts.append(compat_note)
        ws.cell(row=r, column=17, value=" | ".join(note_parts))

        ws.row_dimensions[r].height = DEFAULT_ROW_HEIGHT

    # Conditional formatting for OCI Compatible column (D) — colour updates if cell is edited
    last_data_row = 2 + len(df)
    cf_range = f"D3:D{last_data_row}"
    ws.conditional_formatting.add(cf_range, CellIsRule(operator="equal", formula=['"yes"'],   fill=_CF_FILL_YES))
    ws.conditional_formatting.add(cf_range, CellIsRule(operator="equal", formula=['"maybe"'], fill=_CF_FILL_MAYBE))
    ws.conditional_formatting.add(cf_range, CellIsRule(operator="equal", formula=['"no"'],    fill=_CF_FILL_NO))


def write_os_summary_sheet(wb, vm_df: pd.DataFrame) -> None:
    """Append an 'OS Summary' sheet with COUNTIF formulas that live-update from VM Details."""
    if vm_df is None or vm_df.empty:
        return

    ws = wb.create_sheet(title="OS Summary")

    vd_end       = 2 + len(vm_df)
    vd_os_range  = f"'VM Details'!$C$3:$C${vd_end}"   # OS Detected column
    vd_cf_range  = f"'VM Details'!$D$3:$D${vd_end}"   # OCI Compatible column

    # Column widths
    ws.column_dimensions["A"].width = 37.1
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14

    # ── Row 1: title ──────────────────────────────────────────────────────────
    title_cell = ws.cell(row=1, column=1, value="OS Summary")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws.row_dimensions[1].height = TITLE_ROW_HEIGHT

    cursor = 3  # row 2 intentionally blank

    # ── Section 1: OCI Compatibility Overview ────────────────────────────────
    lbl = ws.cell(row=cursor, column=1, value="OCI Compatibility Overview")
    lbl.font = Font(bold=True)
    lbl.fill = SECTION_LABEL_FILL
    lbl.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=3)
    ws.row_dimensions[cursor].height = DEFAULT_ROW_HEIGHT
    cursor += 1

    for col_idx, hdr in enumerate(["Status", "VM Count", "VM Percent"], start=1):
        cell = ws.cell(row=cursor, column=col_idx, value=hdr)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cursor].height = HEADER_ROW_HEIGHT
    cursor += 1

    compat_data_start = cursor
    compat_total_row  = compat_data_start + 4  # yes/maybe/no/unknown = 4 rows, total is row after

    for status in ("yes", "maybe", "no", "unknown"):
        ws.cell(row=cursor, column=1, value=status)
        ws.cell(row=cursor, column=2, value=f'=COUNTIF({vd_cf_range},"{status}")')
        pct_cell = ws.cell(row=cursor, column=3,
                           value=f"=IF(B${compat_total_row}=0,0,B{cursor}/B${compat_total_row})")
        pct_cell.number_format = "0.0%"
        ws.row_dimensions[cursor].height = DEFAULT_ROW_HEIGHT
        cursor += 1
    compat_data_end = cursor - 1

    total_cell = ws.cell(row=cursor, column=1, value="Total")
    total_cell.font = Font(bold=True)
    count_cell = ws.cell(row=cursor, column=2, value=f"=SUM(B{compat_data_start}:B{compat_data_end})")
    count_cell.font = Font(bold=True)
    ws.row_dimensions[cursor].height = DEFAULT_ROW_HEIGHT
    cursor += 2  # blank gap

    # Conditional formatting on Status column (A) in this section
    cf_status_range = f"A{compat_data_start}:A{compat_data_end}"
    ws.conditional_formatting.add(cf_status_range, CellIsRule(operator="equal", formula=['"yes"'],   fill=_CF_FILL_YES))
    ws.conditional_formatting.add(cf_status_range, CellIsRule(operator="equal", formula=['"maybe"'], fill=_CF_FILL_MAYBE))
    ws.conditional_formatting.add(cf_status_range, CellIsRule(operator="equal", formula=['"no"'],    fill=_CF_FILL_NO))

    # ── Section 2: OS Breakdown ───────────────────────────────────────────────
    lbl2 = ws.cell(row=cursor, column=1, value="OS Breakdown")
    lbl2.font = Font(bold=True)
    lbl2.fill = SECTION_LABEL_FILL
    lbl2.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=cursor, start_column=1, end_row=cursor, end_column=3)
    ws.row_dimensions[cursor].height = DEFAULT_ROW_HEIGHT
    cursor += 1

    for col_idx, hdr in enumerate(["OS Detected", "OCI Compatible", "VM Count"], start=1):
        cell = ws.cell(row=cursor, column=col_idx, value=hdr)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cursor].height = HEADER_ROW_HEIGHT
    cursor += 1

    # Build unique OS list with counts, sorted by count descending then name ascending
    os_data: Dict[str, Tuple[str, int]] = {}
    for _, row in vm_df.iterrows():
        os_str, oci_compat, _ = detect_os(row)
        key = os_str  # empty string = no OS data
        if key not in os_data:
            os_data[key] = (oci_compat, 0)
        os_data[key] = (os_data[key][0], os_data[key][1] + 1)

    sorted_os = sorted(os_data.items(), key=lambda x: (-x[1][1], x[0].lower()))

    os_data_start = cursor
    for os_str, (oci_compat, _) in sorted_os:
        display_os = os_str if os_str else "(No OS data)"
        ws.cell(row=cursor, column=1, value=display_os)
        ws.cell(row=cursor, column=2, value=oci_compat)
        if os_str:
            ws.cell(row=cursor, column=3, value=f"=COUNTIF({vd_os_range},A{cursor})")
        else:
            ws.cell(row=cursor, column=3, value=f'=COUNTIF({vd_os_range},"")')
        ws.row_dimensions[cursor].height = DEFAULT_ROW_HEIGHT
        cursor += 1
    os_data_end = cursor - 1

    total_os = ws.cell(row=cursor, column=1, value="Total")
    total_os.font = Font(bold=True)
    total_os_count = ws.cell(row=cursor, column=3, value=f"=SUM(C{os_data_start}:C{os_data_end})")
    total_os_count.font = Font(bold=True)
    ws.row_dimensions[cursor].height = DEFAULT_ROW_HEIGHT

    # Conditional formatting on OCI Compatible column (B) in OS Breakdown
    cf_os_range = f"B{os_data_start}:B{os_data_end}"
    ws.conditional_formatting.add(cf_os_range, CellIsRule(operator="equal", formula=['"yes"'],   fill=_CF_FILL_YES))
    ws.conditional_formatting.add(cf_os_range, CellIsRule(operator="equal", formula=['"maybe"'], fill=_CF_FILL_MAYBE))
    ws.conditional_formatting.add(cf_os_range, CellIsRule(operator="equal", formula=['"no"'],    fill=_CF_FILL_NO))


def _accounting_number_format(currency: str) -> str:
    """Return an Excel accounting number format string for the given ISO currency code.

    Currencies with a well-known single symbol (USD, EUR, GBP …) use that symbol
    directly.  All others use the [$CODE] notation so Excel displays the ISO code
    left-aligned — identical to what Excel does when you pick Accounting and choose
    the currency from its own dropdown for less common currencies.
    """
    _SYMBOLS: Dict[str, str] = {
        "USD": "$",  "EUR": "€",  "GBP": "£",
        "JPY": "¥",  "CNY": "¥",  "INR": "₹",  "KRW": "₩",
    }
    code = currency.upper()
    sym  = _SYMBOLS.get(code, f"[${code}]")
    return f'_({sym}* #,##0.00_);_({sym}* (#,##0.00);_({sym}* "-"??_);_(@_)'


def write_output(
    output_path: Path,
    metadata: Dict[str, object],
    sheets: Dict[str, List[LineItem]],
    currency: str,
    vm_df: Optional[pd.DataFrame] = None,
) -> None:
    info(f"Writing output to {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    headers = [header.format(currency=currency) for header in EXCEL_HEADERS]
    cost_fmt = _accounting_number_format(currency)

    def roundup_formula(value: float) -> str:
        raw_str = f"{value:.10f}".rstrip("0").rstrip(".")
        if not raw_str:
            raw_str = "0"
        return f"=ROUNDUP({raw_str},0)"

    # Pre-compute vm_details SUM range so total_disk/used_disk can reference it.
    # vm_details data: row 1=headers, row 2=note, row 3..N=VMs.
    _vd_start = 3
    _use_vd_sums = vm_df is not None and not vm_df.empty
    _vd_end = 2 + len(vm_df) if _use_vd_sums else None

    # Column letters in vm_details for each aggregate category
    # (C=OS Detected, D=OCI Compatible inserted before OCPU — so OCPU=E, RAM=F, DiskProv=G, DiskUsed=H)
    _vd_col = {"ocpu": "E", "memory": "F"}  # storage col depends on sheet (G=prov, H=used)

    def _part_qty_formula(category: str, sheet_name: str, raw_quantity: float) -> str:
        """Return a SUM formula referencing vm_details, or ROUNDUP fallback."""
        if not _use_vd_sums:
            return roundup_formula(raw_quantity)
        if category == "storage":
            col = "G" if sheet_name == "total_disk" else "H"
        else:
            col = _vd_col.get(category)
        if col is None:
            return roundup_formula(raw_quantity)
        return f"=SUM('VM Details'!{col}{_vd_start}:{col}{_vd_end})"

    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Summary"

    apply_column_widths(ws)

    # Row 1: title banner
    title_text = f"Oracle Investment Proposal (as of {datetime.now().strftime('%m/%d/%Y')})"
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TABLE_COLUMN_COUNT)
    ws.row_dimensions[1].height = TITLE_ROW_HEIGHT

    # Metadata block (rows 2+) — written once, shared by both sections
    metadata_end_row = write_metadata(ws, build_metadata_rows(metadata), start_row=2)
    for r in range(2, metadata_end_row + 1):
        ws.row_dimensions[r].height = DEFAULT_ROW_HEIGHT
    format_metadata_block(ws, 2, metadata_end_row)

    # Two pricing sections on the same sheet (total_disk then used_disk)
    # Row layout per section:  section_label | table_headers | data×4 | Monthly_Total
    cursor = metadata_end_row + 2   # row 10 = blank gap, row 11 = first section label
    last_total_row = None

    for sheet_key, items in sheets.items():
        # Section label row
        section_label_row = cursor
        lbl = ws.cell(row=section_label_row, column=1, value=SECTION_LABELS.get(sheet_key, sheet_key))
        lbl.font = Font(bold=True)
        lbl.fill = SECTION_LABEL_FILL
        lbl.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(
            start_row=section_label_row, start_column=1,
            end_row=section_label_row, end_column=TABLE_COLUMN_COUNT,
        )
        ws.row_dimensions[section_label_row].height = DEFAULT_ROW_HEIGHT

        # Table headers
        header_row = section_label_row + 1
        format_table_header(ws, header_row, headers)
        ws.row_dimensions[header_row].height = HEADER_ROW_HEIGHT

        # Data rows
        data_start = header_row + 1
        current_row = data_start
        storage_row_index: Optional[int] = None

        for item in items:
            ws.cell(row=current_row, column=1, value=item.description)
            ws.cell(row=current_row, column=2, value=item.part_number)

            part_cell     = ws.cell(row=current_row, column=3)
            instance_cell = ws.cell(row=current_row, column=4)
            usage_cell    = ws.cell(row=current_row, column=5)

            if item.category == "vpu":
                if storage_row_index is None:
                    raise ValueError("Storage row must be written before VPU row.")
                part_cell.value = f"=C{storage_row_index}*B$6"
                instance_cell.value = 1.0
                usage_cell.value = 1.0
            else:
                part_cell.value = _part_qty_formula(item.category, sheet_key, item.raw_base_quantity)
                instance_cell.value = 1.0
                if item.category in {"ocpu", "memory"}:
                    usage_cell.value = "=B$4"
                elif item.category == "storage":
                    usage_cell.value = 1.0
                    storage_row_index = current_row
                else:
                    usage_cell.value = item.usage_quantity

            ws.cell(row=current_row, column=6, value=item.unit_price)
            cost_cell = ws.cell(row=current_row, column=7, value=f"=C{current_row}*D{current_row}*E{current_row}*F{current_row}")
            cost_cell.number_format = cost_fmt
            ws.row_dimensions[current_row].height = DEFAULT_ROW_HEIGHT
            current_row += 1

        total_row = current_row
        ws.cell(row=total_row, column=1, value="Monthly Total")
        for col in range(2, 7):
            ws.cell(row=total_row, column=col, value="")
        total_cost_cell = ws.cell(row=total_row, column=7, value=f"=SUM(G{data_start}:G{current_row - 1})")
        total_cost_cell.number_format = cost_fmt
        ws.row_dimensions[total_row].height = DEFAULT_ROW_HEIGHT

        format_data_rows(ws, data_start, total_row)
        format_total_row(ws, total_row)

        last_total_row = total_row
        cursor = total_row + 2   # blank gap before next section

    # Advisory and disclaimer — written once after both sections
    _, disclaimer_row = append_advisory(ws, last_total_row)
    ws.row_dimensions[disclaimer_row].height = DISCLAIMER_ROW_HEIGHT

    if vm_df is not None and not vm_df.empty:
        write_vm_detail_sheet(
            wb, vm_df, metadata, currency,
            include_vms_off=bool(metadata.get("powered_off_included")),
            include_disks_off=bool(metadata.get("powered_off_disks_included")),
        )
        write_os_summary_sheet(wb, vm_df)

    wb.save(output_path)


# =========================
# CLI / Main
# =========================

def parse_args(argv: Optional[Iterable[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Calculate OCI monthly costs directly from RVTools exports."
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s {VERSION}",
        help="Show program's version number and exit.",
    )
    parser.add_argument(
        "--rvtools",
        nargs="+",
        required=True,
        help="Path(s) to RVTools Excel exports (.xlsx) or directories containing them.",
    )
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Output Excel file path.")
    parser.add_argument("--hours", type=float, default=DEFAULT_HOURS, help="Billing hours in month (default 730).")
    parser.add_argument("--currency", default=DEFAULT_CURRENCY, help="Currency code for pricing (default USD).")
    parser.add_argument("--ocpu-part", default=DEFAULT_OCPU_PART, help="OCI part number for OCPU per hour.")
    parser.add_argument("--memory-part", default=DEFAULT_MEMORY_PART, help="OCI part number for Memory GB per hour.")
    parser.add_argument("--storage-part", default=DEFAULT_STORAGE_PART, help="OCI part number for block storage GB per month.")
    parser.add_argument("--vpu-part", default=DEFAULT_VPU_PART, help="OCI part number for block volume performance units.")
    parser.add_argument("--vpu", type=float, default=DEFAULT_VPU, help="Block volume performance units per GB (default 10).")
    parser.add_argument(
        "--include-poweredoff-vms",
        action="store_true",
        default=False,
        help="Include powered-off VMs in CPU/RAM totals (default: powered-on only).",
    )
    parser.add_argument(
        "--include-poweredoff-disks",
        dest="include_poweredoff_disks",
        action="store_true",
        default=True,
        help="Include powered-off VMs when summing disk usage (default).",
    )
    parser.add_argument(
        "--exclude-poweredoff-disks",
        dest="include_poweredoff_disks",
        action="store_false",
        help="Exclude powered-off VMs when summing disk usage.",
    )
    parser.add_argument(
        "--list",
        action="store_true",
        default=False,
        dest="list_topology",
        help="List all Datacenter and Cluster names found in the input file(s) and exit.",
    )
    parser.add_argument(
        "--datacenter",
        nargs="+",
        metavar="NAME",
        default=None,
        help='Only include VMs in the given Datacenter(s). Quote names with spaces e.g. "DC East".',
    )
    parser.add_argument(
        "--cluster",
        nargs="+",
        metavar="NAME",
        default=None,
        help='Only include VMs in the given Cluster(s). Quote names with spaces e.g. "Cluster East 01".',
    )
    return parser.parse_args(argv)


def main(argv: Optional[Iterable[str]] = None) -> int:
    args = parse_args(argv)

    rvtools_files = collect_rvtools_files(args.rvtools)
    if not rvtools_files:
        print("[ERROR] No RVTools Excel exports found for the given paths.", file=sys.stderr)
        return 1

    if args.list_topology:
        list_datacenters_and_clusters(rvtools_files)
        return 0

    if args.datacenter:
        info(f"Filtering by Datacenter(s): {', '.join(args.datacenter)}")
    if args.cluster:
        info(f"Filtering by Cluster(s): {', '.join(args.cluster)}")

    info(f"Powered-off VMs {'included' if args.include_poweredoff_vms else 'excluded'} in CPU/RAM totals.")
    info(f"Powered-off disks {'included' if args.include_poweredoff_disks else 'excluded'} in disk totals.")

    usage = aggregate_from_rvtools(
        rvtools_files,
        args.include_poweredoff_vms,
        args.include_poweredoff_disks,
        datacenters=args.datacenter,
        clusters=args.cluster,
    )
    if not usage.source_files:
        print("[ERROR] No usable vInfo data found in the provided RVTools exports.", file=sys.stderr)
        return 1

    info(
        f"Aggregated totals - vCPU: {usage.total_vcpu:.2f}, RAM GB: {usage.ram_gb:.2f}, "
        f"Disk Total GB: {usage.disk_total_gb:.2f}, Disk Used GB: {usage.disk_used_gb:.2f}"
    )
    info(f"Powered-on VMs counted: {usage.powered_on_vms}")
    info(f"Powered-off VMs counted: {usage.powered_off_vms}")

    parts = {
        "ocpu": args.ocpu_part,
        "memory": args.memory_part,
        "storage": args.storage_part,
        "vpu": args.vpu_part,
    }

    pricing = PricingClient(args.currency)
    clamped_vpu = max(1.0, min(float(args.vpu), 120.0))

    try:
        sheets = build_line_items(usage, args.hours, clamped_vpu, pricing, parts)
    except Exception as exc:
        print(f"[ERROR] Failed to compute costs: {exc}", file=sys.stderr)
        return 1

    metadata = {
        "source_files": ", ".join(usage.source_files),
        "filter_datacenters": args.datacenter or [],
        "filter_clusters": args.cluster or [],
        "hours_per_month": str(args.hours),
        "currency": args.currency.upper(),
        "vpu": str(clamped_vpu),
        "powered_on_vms": usage.powered_on_vms,
        "powered_off_vms": usage.powered_off_vms,
        "powered_off_included": args.include_poweredoff_vms,
        "powered_off_disks_included": args.include_poweredoff_disks,
    }

    output_path = Path(args.output).resolve()
    try:
        write_output(output_path, metadata, sheets, args.currency.upper(), vm_df=usage.vm_dataframe)
    except Exception as exc:
        print(f"[ERROR] Failed to write output: {exc}", file=sys.stderr)
        return 1

    print("[OK] Cost summary generated successfully.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
